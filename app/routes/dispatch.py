# app/routes/dispatch.py
"""
Módulo de Despacho Kanban

Blueprint para gestión visual de despachos organizados por método de envío.
TODOS los pedidos en estado wc-processing inician en la columna "Por Asignar".
Permite drag & drop de pedidos entre columnas, marcado de prioridades,
y trazabilidad completa de cambios.

Acceso exclusivo para usuario master (Jleon).
"""

from flask import Blueprint, render_template, jsonify, request, current_app
from flask_login import login_required, current_user
from app import db
from app.models import Order, OrderMeta, DispatchHistory, DispatchPriority
from sqlalchemy import text, or_
from datetime import datetime, timedelta
from functools import wraps
import unicodedata
import os
import json

# Crear blueprint
bp = Blueprint('dispatch', __name__, url_prefix='/dispatch')

# Cache para datos de ubigeo (departamentos)
_ubigeo_cache = None


def get_department_name(code):
    """
    Convierte el código del departamento a su nombre completo.

    Args:
        code: Código del departamento (ej: 'AYAC')

    Returns:
        str: Nombre completo del departamento (ej: 'Ayacucho') o el código si no se encuentra
    """
    global _ubigeo_cache

    if not code:
        return None

    # Cargar ubigeo en cache si no está cargado
    if _ubigeo_cache is None:
        try:
            ubigeo_path = os.path.join(current_app.root_path, 'static', 'data', 'ubigeo.json')
            with open(ubigeo_path, 'r', encoding='utf-8') as f:
                _ubigeo_cache = json.load(f)
        except Exception as e:
            current_app.logger.error(f"Error cargando ubigeo.json: {str(e)}")
            return code

    # Buscar el nombre del departamento por código
    for dept in _ubigeo_cache.get('departamentos', []):
        if dept.get('code') == code:
            return dept.get('name', code)

    return code


# ============================================
# MAPEO DE MÉTODOS DE ENVÍO A COLUMNAS
# ============================================

# Mapeo de shipping_method_id (post_id de wpyz_posts) a columnas del Kanban
SHIPPING_METHOD_TO_COLUMN = {
    # SHALOM
    '29884': 'SHALOM',  # Envio Shalom 10
    '15347': 'SHALOM',  # Envio Shalom 12
    '16017': 'SHALOM',  # Zona Maynas Iquitos

    # OLVA COURIER
    '15305': 'Olva Courier',  # Zona Envio Provincia 12.7
    '15369': 'Olva Courier',  # Zona Olva Ica Pisco y Zonas de Lima
    '15310': 'Olva Courier',  # Zona Olva Lima Provincia S/31.90
    '15306': 'Olva Courier',  # Zona Olva Provincia 13.79 13.80
    '15307': 'Olva Courier',  # Zona Olva Provincia 20.7
    '15308': 'Olva Courier',  # Zona Olva Provincia 24.8
    '15311': 'Olva Courier',  # Zona Olva Provincia 32.7 33.70
    '15312': 'Olva Courier',  # Zona Olva Provincia 39.97
    '15304': 'Olva Courier',  # Zona Olva Provincia 9.1
    '15355': 'Olva Courier',  # Zona Olva S/6.90 Lima

    # DINSIDES
    '40672': 'DINSIDES',  # Lima Sur Envio Rapido
    '15302': 'DINSIDES',  # Zona Envio Rapido A
    '15303': 'DINSIDES',  # Zona Envio Rapido B
    '21722': 'DINSIDES',  # Zona Envio Rapido C
    '15300': 'DINSIDES',  # Zona Envio Rapido D
    '15352': 'DINSIDES',  # Zona Rapido L-M-V (1)
    '15354': 'DINSIDES',  # Zona Rapido L-M-V (3)

    # RECOJO EN ALMACÉN
    '15360': 'Recojo en Almacén',  # Zona Recojo
}


def normalize_text(text):
    """
    Normaliza texto removiendo acentos y convirtiéndolo a minúsculas.

    Args:
        text: String a normalizar

    Returns:
        str: Texto normalizado sin acentos en minúsculas
    """
    # Normalizar a NFD (Canonical Decomposition) y remover marcas de acento
    nfd = unicodedata.normalize('NFD', text)
    without_accents = ''.join(char for char in nfd if unicodedata.category(char) != 'Mn')
    return without_accents.lower()


def get_column_from_shipping_method(order_id):
    """
    Determina la columna del Kanban basándose en el método de envío del pedido.

    Args:
        order_id: ID del pedido

    Returns:
        str: Nombre de la columna ('SHALOM', 'Olva Courier', 'DINSIDES', 'Recojo en Almacén',
             'Motorizado (CHAMO)', o 'Por Asignar')
    """
    try:
        # Obtener el nombre del método de envío (con fallback para pedidos sin shipping item)
        query = text("""
            SELECT
                COALESCE(
                    (SELECT oi.order_item_name
                     FROM wpyz_woocommerce_order_items oi
                     WHERE oi.order_id = o.id
                       AND oi.order_item_type = 'shipping'
                     LIMIT 1),
                    -- Fallback: extraer desde _billing_entrega metadata
                    CASE
                        WHEN om_billing_entrega.meta_value = 'billing_recojo' THEN 'Recojo en Almacén'
                        WHEN om_billing_entrega.meta_value LIKE '%recojo%' THEN 'Recojo en Almacén'
                        WHEN om_billing_entrega.meta_value = 'billing_address' THEN 'Envío a domicilio'
                        ELSE NULL
                    END
                ) as shipping_method
            FROM wpyz_wc_orders o
            LEFT JOIN wpyz_wc_orders_meta om_billing_entrega
                ON o.id = om_billing_entrega.order_id
                AND om_billing_entrega.meta_key = '_billing_entrega'
            WHERE o.id = :order_id
        """)

        result = db.session.execute(query, {'order_id': order_id}).fetchone()

        if result and result[0]:
            # Normalizar el nombre del método de envío (quitar acentos y minúsculas)
            shipping_method_name = normalize_text(result[0])

            # Mapeo por palabras clave en el nombre del método (sin acentos)
            # SHALOM
            if 'shalom' in shipping_method_name or 'maynas' in shipping_method_name or 'iquitos' in shipping_method_name:
                return 'SHALOM'

            # OLVA COURIER
            elif 'olva' in shipping_method_name:
                return 'Olva Courier'

            # RECOJO EN ALMACÉN
            elif 'recojo' in shipping_method_name:
                return 'Recojo en Almacén'

            # DINSIDES (Envío rápido, Lima Sur, 1 día hábil, etc.)
            # Búsqueda sin acentos: "envio 1 dia habil" coincide con "Envío 1 día hábil"
            elif any(keyword in shipping_method_name for keyword in ['rapido', 'lima sur', 'envio rapido', '1 dia', 'dia habil']):
                return 'DINSIDES'

            else:
                current_app.logger.warning(f"[DISPATCH] Order {order_id}: Nombre '{result[0]}' NO coincide con ningún patrón (normalizado: '{shipping_method_name}')")

        else:
            current_app.logger.warning(f"[DISPATCH] Order {order_id}: NO tiene método de envío")

        # Si no se encuentra mapeo, va a "Por Asignar"
        current_app.logger.info(f"[DISPATCH] Order {order_id}: Asignado a 'Por Asignar' (sin mapeo)")
        return 'Por Asignar'

    except Exception as e:
        current_app.logger.error(f"[DISPATCH] Error obteniendo columna para pedido {order_id}: {str(e)}")
        return 'Por Asignar'


# ============================================
# MIDDLEWARE DE AUTORIZACIÓN
# ============================================

def master_required(f):
    """
    Decorador para restringir acceso solo a usuarios con role 'master'

    Uso:
        @bp.route('/ruta')
        @login_required
        @master_required
        def mi_funcion():
            ...
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return jsonify({
                'success': False,
                'error': 'No autenticado'
            }), 401

        if current_user.role != 'master':
            current_app.logger.warning(
                f"Usuario {current_user.username} (role: {current_user.role}) "
                f"intentó acceder a módulo de despacho sin permisos"
            )
            return jsonify({
                'success': False,
                'error': 'Acceso denegado. Este módulo es exclusivo para usuarios master.'
            }), 403

        return f(*args, **kwargs)

    return decorated_function


# ============================================
# RUTAS PRINCIPALES
# ============================================

@bp.route('/')
@login_required
@master_required
def index():
    """
    Vista principal del tablero Kanban de despacho

    Muestra pedidos organizados por método de envío en columnas:
    - Por Asignar (TODOS los pedidos wc-processing inician aquí)
    - Olva Courier
    - Recojo en Almacén
    - Motorizado (CHAMO)
    - SHALOM
    - DINSIDES
    """
    current_app.logger.info(f"Usuario {current_user.username} accedió al módulo de despacho")

    return render_template(
        'dispatch_board.html',
        page_title='Módulo de Despacho',
        user=current_user
    )


# ============================================
# API ENDPOINTS
# ============================================

@bp.route('/api/debug', methods=['GET'])
@login_required
@master_required
def debug_params():
    """
    Endpoint temporal de debug para ver qué parámetros llegan
    """
    return jsonify({
        'success': True,
        'debug': {
            'all_args': dict(request.args),
            'date_from': request.args.get('date_from'),
            'date_to': request.args.get('date_to'),
            'priority_only': request.args.get('priority_only'),
            'raw_query_string': request.query_string.decode('utf-8')
        }
    })


@bp.route('/api/orders', methods=['GET'])
@login_required
@master_required
def get_orders():
    """
    Obtener pedidos agrupados por método de envío

    Query Parameters:
        - date_from: Fecha inicio (YYYY-MM-DD)
        - date_to: Fecha fin (YYYY-MM-DD)
        - priority_only: Si es 'true', solo pedidos prioritarios
        - shipping_methods: Métodos de envío separados por coma

    Returns:
        JSON con pedidos agrupados por método de envío
    """
    try:
        # Parámetros de filtro
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        priority_only = request.args.get('priority_only', 'false').lower() == 'true'
        atendido_only = request.args.get('atendido_only', 'false').lower() == 'true'
        no_atendido_only = request.args.get('no_atendido_only', 'false').lower() == 'true'
        shipping_methods_filter = request.args.get('shipping_methods', '').split(',') if request.args.get('shipping_methods') else None

        # Convertir formato de fecha si viene en formato dd/mm/yyyy a yyyy-mm-dd
        if date_from and '/' in date_from:
            # Formato dd/mm/yyyy -> yyyy-mm-dd
            parts = date_from.split('/')
            if len(parts) == 3:
                date_from = f"{parts[2]}-{parts[1]}-{parts[0]}"

        if date_to and '/' in date_to:
            # Formato dd/mm/yyyy -> yyyy-mm-dd
            parts = date_to.split('/')
            if len(parts) == 3:
                date_to = f"{parts[2]}-{parts[1]}-{parts[0]}"

        # Query base: TODOS los pedidos wc-processing (WooCommerce nativos + WhatsApp)
        query = text("""
            SELECT DISTINCT
                o.id,
                COALESCE(om_number.meta_value, CONCAT('#', o.id)) as order_number,
                om_number.meta_value as whatsapp_number,  -- Número W-XXXXX si existe
                DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR) as date_created_local,
                o.total_amount,
                o.status,
                o.billing_email,

                -- Datos de cliente
                ba.first_name,
                ba.last_name,
                ba.phone,

                -- Método de envío (con fallback para pedidos sin shipping item)
                COALESCE(
                    (SELECT oi.order_item_name
                     FROM wpyz_woocommerce_order_items oi
                     WHERE oi.order_id = o.id
                       AND oi.order_item_type = 'shipping'
                     LIMIT 1),
                    -- Fallback: extraer desde _billing_entrega metadata
                    CASE
                        WHEN om_billing_entrega.meta_value = 'billing_recojo' THEN 'Recojo en Almacén'
                        WHEN om_billing_entrega.meta_value LIKE '%recojo%' THEN 'Recojo en Almacén'
                        WHEN om_billing_entrega.meta_value = 'billing_address' THEN 'Envío a domicilio'
                        ELSE NULL
                    END
                ) as shipping_method,

                -- Prioridad (si existe)
                dp.is_priority,
                dp.priority_level,
                dp.is_atendido,

                -- Tiempo sin mover (horas desde última actualización)
                TIMESTAMPDIFF(HOUR, o.date_updated_gmt, UTC_TIMESTAMP()) as hours_since_update,

                -- Usuario que creó el pedido
                om_created.meta_value as created_by,

                -- Distrito de envío (para pedidos 1 día hábil)
                sa.city as shipping_district

            FROM wpyz_wc_orders o

            -- Número de pedido (opcional - puede no tener W-XXXXX si es WooCommerce nativo)
            LEFT JOIN wpyz_wc_orders_meta om_number
                ON o.id = om_number.order_id
                AND om_number.meta_key = '_order_number'

            -- Dirección de facturación (datos de cliente)
            LEFT JOIN wpyz_wc_order_addresses ba
                ON o.id = ba.order_id
                AND ba.address_type = 'billing'

            -- Dirección de envío (para obtener distrito)
            LEFT JOIN wpyz_wc_order_addresses sa
                ON o.id = sa.order_id
                AND sa.address_type = 'shipping'

            -- Prioridades
            LEFT JOIN woo_dispatch_priorities dp
                ON o.id = dp.order_id

            -- Usuario que creó
            LEFT JOIN wpyz_wc_orders_meta om_created
                ON o.id = om_created.order_id
                AND om_created.meta_key = '_created_by'

            -- Metadata de tipo de entrega (fallback para shipping_method)
            LEFT JOIN wpyz_wc_orders_meta om_billing_entrega
                ON o.id = om_billing_entrega.order_id
                AND om_billing_entrega.meta_key = '_billing_entrega'

            WHERE o.status = 'wc-processing'

            -- Filtro por fecha
            {date_filter}

            -- Filtro por prioridad
            {priority_filter}

            -- Filtro por estado de atendido
            {atendido_filter}

            ORDER BY
                dp.is_priority DESC,
                dp.priority_level DESC,
                hours_since_update DESC  -- Pedidos con más retraso primero
        """)

        # Construir filtros dinámicos
        date_filter = ""
        priority_filter = ""
        atendido_filter = ""
        params = {}

        # NOTA: El filtro de fechas es OPCIONAL
        # Por defecto, muestra TODOS los pedidos en estado wc-processing
        # Solo filtra por fecha si el usuario aplica el filtro manualmente
        # IMPORTANTE: Usar DATE_SUB para convertir GMT a hora de Perú (UTC-5)
        if date_from and date_to:
            date_filter = "AND DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR)) BETWEEN :date_from AND :date_to"
            params['date_from'] = date_from
            params['date_to'] = date_to

        if priority_only:
            priority_filter = "AND dp.is_priority = TRUE"

        # Filtros de estado de atendido (mutuamente excluyentes)
        if atendido_only:
            atendido_filter = "AND dp.is_atendido = TRUE"
        elif no_atendido_only:
            atendido_filter = "AND (dp.is_atendido = FALSE OR dp.is_atendido IS NULL)"

        # Reemplazar placeholders
        query_str = str(query).format(
            date_filter=date_filter,
            priority_filter=priority_filter,
            atendido_filter=atendido_filter
        )

        # Log para debug
        current_app.logger.info("="*80)
        current_app.logger.info(f"DISPATCH API - GET ORDERS REQUEST")
        current_app.logger.info(f"date_from (convertido): {date_from}")
        current_app.logger.info(f"date_to (convertido): {date_to}")
        current_app.logger.info(f"priority_only: {priority_only}")
        current_app.logger.info(f"Query params: {params}")
        current_app.logger.info(f"Date filter SQL: {date_filter}")
        current_app.logger.info("="*80)

        # Ejecutar query
        results = db.session.execute(text(query_str), params).fetchall()

        current_app.logger.info(f"✓ Pedidos encontrados: {len(results)}")
        if len(results) > 0:
            current_app.logger.info(f"  Primer pedido: ID={results[0][0]}, Numero={results[0][1]}, Fecha={results[0][2]}")
        current_app.logger.info("="*80)

        # Agrupar por método de envío
        orders_by_method = {
            'Por Asignar': [],
            'Olva Courier': [],
            'Recojo en Almacén': [],
            'Motorizado (CHAMO)': [],
            'SHALOM': [],
            'DINSIDES': []
        }

        # Obtener última ubicación de cada pedido desde el historial
        order_ids = [row[0] for row in results]
        last_positions = {}

        if order_ids:
            # Query para obtener el último movimiento de cada pedido
            positions_query = text("""
                SELECT
                    dh.order_id,
                    dh.new_shipping_method
                FROM woo_dispatch_history dh
                INNER JOIN (
                    SELECT order_id, MAX(changed_at) as last_changed
                    FROM woo_dispatch_history
                    WHERE order_id IN :order_ids
                    GROUP BY order_id
                ) latest ON dh.order_id = latest.order_id
                    AND dh.changed_at = latest.last_changed
            """)

            positions_result = db.session.execute(
                positions_query,
                {'order_ids': tuple(order_ids)}
            ).fetchall()

            # Mapear order_id -> última columna
            for pos_row in positions_result:
                last_positions[pos_row[0]] = pos_row[1]

        for row in results:
            order_id = row[0]

            # Verificar si el pedido tiene un movimiento previo en el historial
            if order_id in last_positions:
                # Si tiene historial, usar la última posición guardada
                column = last_positions[order_id]
            else:
                # Si no hay historial, determinar columna automáticamente según método de envío
                column = get_column_from_shipping_method(order_id)

            # Aplicar filtro de métodos si existe
            if shipping_methods_filter and column not in shipping_methods_filter:
                continue

            # Construir objeto de pedido
            whatsapp_number = row[2]  # W-XXXXX si existe, None si no
            display_number = row[1]   # Ya viene como COALESCE(W-XXXXX, #ID)

            # Si es pedido WhatsApp, mostrar #ID principal y W-XXXXX secundario
            if whatsapp_number and whatsapp_number.startswith('W-'):
                display_number = f"#{row[0]}"  # ID original
                whatsapp_label = whatsapp_number  # W-XXXXX
            else:
                whatsapp_label = None

            order_data = {
                'id': row[0],
                'number': display_number,
                'whatsapp_number': whatsapp_label,  # W-XXXXX para mostrar en gris
                'date_created': row[3].strftime('%Y-%m-%d %H:%M') if row[3] else None,
                'total': float(row[4]) if row[4] else 0,
                'status': row[5],
                'email': row[6],
                'customer_name': f"{row[7]} {row[8]}" if row[7] and row[8] else 'N/A',
                'customer_phone': row[9] or 'N/A',
                'shipping_method': row[10] or 'Sin método',
                'is_priority': bool(row[11]) if row[11] is not None else False,
                'priority_level': row[12] or 'normal',
                'is_atendido': bool(row[13]) if row[13] is not None else False,
                'hours_since_update': row[14] or 0,
                'is_stale': (row[14] or 0) > 24,  # Más de 24h sin mover
                'created_by': row[15] or 'Desconocido',
                'shipping_district': row[16] or None  # Distrito de envío
            }

            # Agregar a la columna correspondiente (con fallback de seguridad)
            orders_by_method.get(column, orders_by_method['Por Asignar']).append(order_data)

        # Calcular estadísticas
        total_orders = sum(len(orders) for orders in orders_by_method.values())
        priority_orders = sum(
            1 for orders in orders_by_method.values()
            for order in orders
            if order['is_priority']
        )
        atendido_orders = sum(
            1 for orders in orders_by_method.values()
            for order in orders
            if order['is_atendido']
        )
        stale_orders = sum(
            1 for orders in orders_by_method.values()
            for order in orders
            if order['is_stale']
        )

        return jsonify({
            'success': True,
            'orders': orders_by_method,
            'stats': {
                'total': total_orders,
                'priority': priority_orders,
                'atendido': atendido_orders,
                'stale': stale_orders
            }
        })

    except Exception as e:
        current_app.logger.error(f"Error obteniendo pedidos para despacho: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/api/move', methods=['POST'])
@login_required
@master_required
def move_order():
    """
    Mover pedido a otra columna (cambiar método de envío)

    Request Body:
        {
            "order_id": 123,
            "new_shipping_method": "Olva Courier"
        }

    Returns:
        JSON con confirmación del cambio
    """
    try:
        data = request.get_json()
        order_id = data.get('order_id')
        new_shipping_method = data.get('new_shipping_method')

        if not order_id or not new_shipping_method:
            return jsonify({
                'success': False,
                'error': 'Parámetros faltantes: order_id y new_shipping_method son requeridos'
            }), 400

        # Obtener pedido
        order = Order.query.get(order_id)
        if not order:
            return jsonify({
                'success': False,
                'error': f'Pedido {order_id} no encontrado'
            }), 404

        # Obtener número de pedido (puede ser None para pedidos WooCommerce nativos)
        order_number = order.get_meta('_order_number')
        if not order_number:
            order_number = f"#{order_id}"

        # Obtener método de envío actual (ORIGINAL del pedido)
        current_shipping = db.session.execute(
            text("""
                SELECT order_item_name
                FROM wpyz_woocommerce_order_items
                WHERE order_id = :order_id
                  AND order_item_type = 'shipping'
                LIMIT 1
            """),
            {'order_id': order_id}
        ).scalar()

        # NO actualizar el método de envío en la base de datos
        # Solo registrar el movimiento en el historial de despacho
        # Esto permite mantener el método de envío original del pedido

        # Registrar en historial
        history_entry = DispatchHistory(
            order_id=order_id,
            order_number=order_number,
            previous_shipping_method=current_shipping,
            new_shipping_method=new_shipping_method,
            changed_by=current_user.username,
            changed_at=datetime.utcnow()
        )
        db.session.add(history_entry)

        # Actualizar date_updated_gmt del pedido
        order.date_updated_gmt = datetime.utcnow()

        db.session.commit()

        current_app.logger.info(
            f"Pedido {order_number} movido de '{current_shipping}' a '{new_shipping_method}' "
            f"por {current_user.username}"
        )

        return jsonify({
            'success': True,
            'message': f'Pedido {order_number} movido exitosamente',
            'order_id': order_id,
            'order_number': order_number,
            'previous_method': current_shipping,
            'new_method': new_shipping_method
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error moviendo pedido: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/api/priority', methods=['POST'])
@login_required
@master_required
def set_priority():
    """
    Marcar/desmarcar pedido como prioritario

    Request Body:
        {
            "order_id": 123,
            "is_priority": true,
            "priority_level": "urgent",  // 'normal', 'high', 'urgent'
            "note": "Cliente requiere envío urgente"
        }

    Returns:
        JSON con confirmación
    """
    try:
        data = request.get_json()
        order_id = data.get('order_id')
        is_priority = data.get('is_priority', False)
        priority_level = data.get('priority_level', 'normal')
        note = data.get('note', '')

        if not order_id:
            return jsonify({
                'success': False,
                'error': 'order_id es requerido'
            }), 400

        # Obtener pedido
        order = Order.query.get(order_id)
        if not order:
            return jsonify({
                'success': False,
                'error': f'Pedido {order_id} no encontrado'
            }), 404

        order_number = order.get_meta('_order_number')
        if not order_number:
            order_number = f"#{order_id}"

        # Buscar o crear registro de prioridad
        priority = DispatchPriority.query.filter_by(order_id=order_id).first()

        if not priority:
            priority = DispatchPriority(
                order_id=order_id,
                order_number=order_number
            )
            db.session.add(priority)

        # Actualizar valores
        priority.is_priority = is_priority
        priority.priority_level = priority_level if is_priority else 'normal'
        priority.marked_by = current_user.username if is_priority else None
        priority.marked_at = datetime.utcnow() if is_priority else None
        priority.priority_note = note if is_priority else None

        db.session.commit()

        action = "marcado como prioritario" if is_priority else "desmarcado como prioritario"
        current_app.logger.info(
            f"Pedido {order_number} {action} (nivel: {priority_level}) "
            f"por {current_user.username}"
        )

        return jsonify({
            'success': True,
            'message': f'Pedido {order_number} {action}',
            'order_id': order_id,
            'order_number': order_number,
            'is_priority': is_priority,
            'priority_level': priority_level
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error configurando prioridad: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/api/atendido', methods=['POST'])
@login_required
@master_required
def set_atendido():
    """
    Marcar/desmarcar pedido como atendido/empaquetado (Listo para envío)

    Request Body:
        {
            "order_id": 123,
            "is_atendido": true
        }

    Returns:
        JSON con confirmación
    """
    try:
        data = request.get_json()
        order_id = data.get('order_id')
        is_atendido = data.get('is_atendido', False)

        if not order_id:
            return jsonify({
                'success': False,
                'error': 'order_id es requerido'
            }), 400

        # Obtener pedido
        order = Order.query.get(order_id)
        if not order:
            return jsonify({
                'success': False,
                'error': f'Pedido {order_id} no encontrado'
            }), 404

        order_number = order.get_meta('_order_number')
        if not order_number:
            order_number = f"#{order_id}"

        # Buscar o crear registro de prioridad (donde guardamos el estado atendido)
        priority = DispatchPriority.query.filter_by(order_id=order_id).first()

        if not priority:
            priority = DispatchPriority(
                order_id=order_id,
                order_number=order_number
            )
            db.session.add(priority)

        # Actualizar valores
        priority.is_atendido = is_atendido
        priority.atendido_by = current_user.username if is_atendido else None
        priority.atendido_at = datetime.utcnow() if is_atendido else None

        db.session.commit()

        action = "marcado como atendido/empaquetado" if is_atendido else "desmarcado como atendido"
        current_app.logger.info(
            f"Pedido {order_number} {action} por {current_user.username}"
        )

        return jsonify({
            'success': True,
            'message': f'Pedido {order_number} {action}',
            'order_id': order_id,
            'order_number': order_number,
            'is_atendido': is_atendido
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error configurando estado atendido: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/api/note', methods=['POST'])
@login_required
@master_required
def add_note():
    """
    Agregar nota de despacho a un pedido

    Request Body:
        {
            "order_id": 123,
            "note": "Empacar con cuidado - producto frágil"
        }

    Returns:
        JSON con confirmación
    """
    try:
        data = request.get_json()
        order_id = data.get('order_id')
        note = data.get('note', '').strip()

        if not order_id or not note:
            return jsonify({
                'success': False,
                'error': 'order_id y note son requeridos'
            }), 400

        # Obtener pedido
        order = Order.query.get(order_id)
        if not order:
            return jsonify({
                'success': False,
                'error': f'Pedido {order_id} no encontrado'
            }), 404

        order_number = order.get_meta('_order_number')

        # Obtener la columna actual del tablero para este pedido
        # (Usa el mapeo centralizado en lugar del nombre crudo del método)
        current_column = get_column_from_shipping_method(order_id)

        # Crear entrada en historial con solo nota (sin cambio de método de despacho)
        history_entry = DispatchHistory(
            order_id=order_id,
            order_number=order_number,
            previous_shipping_method=None,
            new_shipping_method=current_column,
            changed_by=current_user.username,
            changed_at=datetime.utcnow(),
            dispatch_note=note
        )
        db.session.add(history_entry)
        db.session.commit()

        current_app.logger.info(
            f"Nota agregada al pedido {order_number} por {current_user.username}: {note[:50]}..."
        )

        return jsonify({
            'success': True,
            'message': 'Nota agregada exitosamente',
            'order_id': order_id,
            'order_number': order_number,
            'note': note
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error agregando nota: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/api/history/<int:order_id>', methods=['GET'])
@login_required
@master_required
def get_history(order_id):
    """
    Obtener historial completo de cambios de un pedido

    Returns:
        JSON con lista de cambios ordenados cronológicamente
    """
    try:
        history = DispatchHistory.query.filter_by(order_id=order_id)\
            .order_by(DispatchHistory.changed_at.desc())\
            .all()

        return jsonify({
            'success': True,
            'order_id': order_id,
            'history': [entry.to_dict() for entry in history]
        })

    except Exception as e:
        current_app.logger.error(f"Error obteniendo historial: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/api/order/<int:order_id>', methods=['GET'])
@login_required
@master_required
def get_order_detail(order_id):
    """
    Obtener detalles completos de un pedido incluyendo productos

    Returns:
        JSON con información completa del pedido y lista de productos
    """
    try:
        # Query para obtener detalles del pedido
        order_query = text("""
            SELECT
                o.id,
                COALESCE(om_number.meta_value, CONCAT('#', o.id)) as order_number,
                o.date_created_gmt,
                o.total_amount,
                o.status,
                o.billing_email,
                ba.first_name,
                ba.last_name,
                ba.phone,
                (SELECT oi.order_item_name
                 FROM wpyz_woocommerce_order_items oi
                 WHERE oi.order_id = o.id
                   AND oi.order_item_type = 'shipping'
                 LIMIT 1) as shipping_method,
                om_created.meta_value as created_by,
                -- Dirección de envío
                sa.address_1 as shipping_address,
                sa.address_2 as shipping_address_2,
                sa.city as shipping_district,
                sa.state as shipping_department,
                sa.postcode as shipping_postcode,
                -- Notas del cliente
                o.customer_note,
                -- DNI del cliente (guardado en billing company)
                ba.company as customer_dni
            FROM wpyz_wc_orders o
            LEFT JOIN wpyz_wc_orders_meta om_number
                ON o.id = om_number.order_id
                AND om_number.meta_key = '_order_number'
            LEFT JOIN wpyz_wc_order_addresses ba
                ON o.id = ba.order_id
                AND ba.address_type = 'billing'
            LEFT JOIN wpyz_wc_order_addresses sa
                ON o.id = sa.order_id
                AND sa.address_type = 'shipping'
            LEFT JOIN wpyz_wc_orders_meta om_created
                ON o.id = om_created.order_id
                AND om_created.meta_key = '_created_by'
            WHERE o.id = :order_id
        """)

        order_result = db.session.execute(order_query, {'order_id': order_id}).fetchone()

        if not order_result:
            return jsonify({
                'success': False,
                'error': f'Pedido {order_id} no encontrado'
            }), 404

        # Query para obtener productos del pedido con SKU
        # NOTA: Los atributos los extraemos del order_item_name directamente
        products_query = text("""
            SELECT
                oi.order_item_name as product_name,
                oim_qty.meta_value as quantity,
                COALESCE(oim_variation_id.meta_value, oim_product_id.meta_value) as product_id,
                (SELECT pm.meta_value
                 FROM wpyz_postmeta pm
                 WHERE pm.post_id = COALESCE(oim_variation_id.meta_value, oim_product_id.meta_value)
                   AND pm.meta_key = '_sku'
                 LIMIT 1) as sku
            FROM wpyz_woocommerce_order_items oi
            LEFT JOIN wpyz_woocommerce_order_itemmeta oim_qty
                ON oi.order_item_id = oim_qty.order_item_id
                AND oim_qty.meta_key = '_qty'
            LEFT JOIN wpyz_woocommerce_order_itemmeta oim_product_id
                ON oi.order_item_id = oim_product_id.order_item_id
                AND oim_product_id.meta_key = '_product_id'
            LEFT JOIN wpyz_woocommerce_order_itemmeta oim_variation_id
                ON oi.order_item_id = oim_variation_id.order_item_id
                AND oim_variation_id.meta_key = '_variation_id'
                AND oim_variation_id.meta_value != '0'
            WHERE oi.order_id = :order_id
              AND oi.order_item_type = 'line_item'
            ORDER BY oi.order_item_id
        """)

        products_result = db.session.execute(products_query, {'order_id': order_id}).fetchall()

        # Construir lista de productos con imágenes
        products_list = []
        for row in products_result:
            product_id = row[2]
            product_name_raw = row[0]

            # DEBUG: Log para ver qué está trayendo la BD
            current_app.logger.info(f"DEBUG - Product name raw from DB: {product_name_raw}")

            # Limpiar el nombre del producto - eliminar basura técnica de WooCommerce
            # Formato sucio: "Nombre - Atributos | line_subtotal_tax: ... | reduced_stock: ..."
            # Queremos solo: "Nombre - Atributos"
            product_name_clean = product_name_raw
            if ' | line' in product_name_raw or ' | reduced' in product_name_raw or ' | tax' in product_name_raw:
                # Eliminar todo después del primer " | " que contiene datos técnicos
                product_name_clean = product_name_raw.split(' | line')[0].split(' | reduced')[0].split(' | tax')[0].strip()

            # Extraer atributos del nombre limpio del producto
            # Formato: "Nombre - Atributo1, Atributo2"
            # Los atributos están después del último " - " o después de la primera coma
            attributes_from_name = None
            clean_product_name = product_name_clean

            if ' - ' in product_name_clean:
                # Dividir por el último " - "
                parts = product_name_clean.rsplit(' - ', 1)
                if len(parts) == 2:
                    clean_product_name = parts[0].strip()
                    # La segunda parte contiene los atributos
                    attributes_from_name = parts[1].strip()
            elif ',' in product_name_clean and ' | ' not in product_name_clean:
                # Si no hay " - " pero hay coma (y no es basura técnica), todo después de la coma son atributos
                parts = product_name_clean.split(',', 1)
                if len(parts) == 2:
                    clean_product_name = parts[0].strip()
                    attributes_from_name = parts[1].strip()

            # Los atributos solo vienen del nombre (row[4] ya no existe - eliminamos el GROUP_CONCAT)
            final_attributes = attributes_from_name

            # Obtener imagen del producto
            image_url = None
            if product_id:
                image_query = text("""
                    SELECT pm.meta_value as image_id
                    FROM wpyz_postmeta pm
                    WHERE pm.post_id = :product_id
                      AND pm.meta_key = '_thumbnail_id'
                    LIMIT 1
                """)
                image_result = db.session.execute(image_query, {'product_id': product_id}).fetchone()

                if image_result and image_result[0]:
                    image_id = image_result[0]
                    # Obtener URL de la imagen
                    image_url_query = text("""
                        SELECT guid
                        FROM wpyz_posts
                        WHERE ID = :image_id
                    """)
                    image_url_result = db.session.execute(image_url_query, {'image_id': image_id}).fetchone()
                    if image_url_result:
                        image_url = image_url_result[0]

            products_list.append({
                'name': clean_product_name,
                'quantity': int(row[1]) if row[1] else 1,
                'sku': row[3] if row[3] else None,
                'attributes': final_attributes,
                'image': image_url
            })

        # Construir dirección completa
        address_parts = []
        if order_result[11]:  # shipping_address
            address_parts.append(order_result[11])
        if order_result[12]:  # shipping_address_2
            address_parts.append(order_result[12])
        shipping_address_full = ', '.join(address_parts) if address_parts else None

        # Construir respuesta
        order_data = {
            'id': order_result[0],
            'number': order_result[1],
            'date_created': order_result[2].strftime('%Y-%m-%d %H:%M') if order_result[2] else None,
            'total': float(order_result[3]) if order_result[3] else 0,
            'status': order_result[4],
            'email': order_result[5],
            'customer_name': f"{order_result[6]} {order_result[7]}" if order_result[6] and order_result[7] else 'N/A',
            'customer_phone': order_result[8] or 'N/A',
            'shipping_method': order_result[9] or 'Sin método',
            'created_by': order_result[10] or 'Desconocido',
            # Nuevos campos de dirección
            'shipping_address': shipping_address_full,
            'shipping_district': order_result[13] or None,  # city = distrito
            'shipping_department': get_department_name(order_result[14]),  # state = departamento (convertido a nombre)
            'shipping_postcode': order_result[15] or None,
            # Notas del cliente
            'customer_note': order_result[16] or None,
            # DNI del cliente
            'customer_dni': order_result[17] or None,
            'products': products_list
        }

        return jsonify({
            'success': True,
            'order': order_data
        })

    except Exception as e:
        current_app.logger.error(f"Error obteniendo detalle de pedido: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/api/add-tracking', methods=['POST'])
@login_required
@master_required
def add_tracking():
    """
    Agregar información de tracking a un pedido.

    Request Body:
        {
            "order_id": 41608,
            "tracking_number": "IZI26010841608660",
            "shipping_provider": "Motorizado Izi",
            "date_shipped": "2026-01-08",  // Opcional, por defecto fecha actual
            "mark_as_shipped": true  // Opcional, si cambiar a wc-completed
        }

    Returns:
        JSON con confirmación del cambio
    """
    try:
        data = request.get_json()
        order_id = data.get('order_id')
        tracking_number = data.get('tracking_number')
        shipping_provider = data.get('shipping_provider')
        date_shipped = data.get('date_shipped')
        mark_as_shipped = data.get('mark_as_shipped', False)

        # Validar parámetros requeridos
        if not all([order_id, tracking_number, shipping_provider]):
            return jsonify({
                'success': False,
                'error': 'Parámetros faltantes: order_id, tracking_number y shipping_provider son requeridos'
            }), 400

        # Validar que el pedido existe
        order = Order.query.get(order_id)
        if not order:
            return jsonify({
                'success': False,
                'error': f'Pedido {order_id} no encontrado'
            }), 404

        # Fecha de envío (por defecto hoy)
        if not date_shipped:
            from datetime import date
            date_shipped = date.today().strftime('%Y-%m-%d')

        # Timestamp actual
        timestamp = int(datetime.utcnow().timestamp())

        # Crear el array de tracking items en formato PHP serializado
        # Estructura: array con un item que contiene tracking_number, tracking_provider, date_shipped
        tracking_items = [{
            'tracking_number': tracking_number,
            'tracking_provider': shipping_provider,
            'custom_tracking_provider': '',
            'custom_tracking_link': '',
            'date_shipped': date_shipped,
            'tracking_id': str(timestamp)  # ID único basado en timestamp
        }]

        # Serializar a formato PHP
        import phpserialize
        serialized_items = phpserialize.dumps(tracking_items).decode('utf-8')

        current_app.logger.info(f"[TRACKING] Order {order_id}: Agregando tracking {tracking_number} con provider '{shipping_provider}'")
        current_app.logger.info(f"[TRACKING] Serialized data: {serialized_items}")

        # Guardar metadatos en wpyz_postmeta (para compatibilidad con el plugin)
        # IMPORTANTE: El plugin de WooCommerce Shipment Tracking requiere registros DUPLICADOS
        # Por eso insertamos cada meta_key DOS VECES (sin ON DUPLICATE KEY UPDATE)

        # 1. ACTUALIZAR VÍA API PRIMERO (si se marca como enviado)
        # Esto triggera los emails y el cambio de estado oficial en WooCommerce
        if mark_as_shipped:
            try:
                from woocommerce import API
                wc_api = API(
                    url=current_app.config['WC_API_URL'],
                    consumer_key=current_app.config['WC_CONSUMER_KEY'],
                    consumer_secret=current_app.config['WC_CONSUMER_SECRET'],
                    version="wc/v3",
                    timeout=15 
                )

                # Preservar payment_method del pedido
                payment_method = order.get_meta('_payment_method') or order.payment_method

                # IMPORTANTE: Enviamos tracking_items (LISTA), NO serialized_items (STRING)
                # La API de WooCommerce se encarga de serializarlo CORRECTAMENTE (una sola vez)
                wc_api.put(f"orders/{order_id}", {
                    "status": "completed",
                    "payment_method": payment_method,
                    "meta_data": [
                        {"key": "_tracking_number", "value": tracking_number},
                        {"key": "_tracking_provider", "value": shipping_provider},
                        {"key": "_wc_shipment_tracking_items", "value": tracking_items}
                    ]
                })
                current_app.logger.info(f"[TRACKING] Order {order_id}: API update success (status: completed)")
                
                # Actualizar objeto local
                order.status = 'wc-completed'
                order.date_updated_gmt = datetime.utcnow()

            except Exception as api_err:
                current_app.logger.error(f"[TRACKING] API Error (non-blocking): {str(api_err)}")
                # Si la API falla, intentamos al menos cambiar el estado localmente
                order.status = 'wc-completed'
                order.date_updated_gmt = datetime.utcnow()

        # 2. GUARDAR EN BASE DE DATOS (DUPLICADOS PARA EL PLUGIN)
        # Hacemos esto DESPUÉS de la API (o independientemente) para asegurar que 
        # los registros duplicados persistan en el modelo Legacy de WooCommerce.
        
        # Eliminar registros anteriores para evitar acumulación
        query_delete_old = text("""
            DELETE FROM wpyz_postmeta 
            WHERE post_id = :order_id 
              AND meta_key IN ('_tracking_number', '_tracking_provider', '_wc_shipment_tracking_items')
        """)
        db.session.execute(query_delete_old, {'order_id': order_id})

        # Insertar DUPLICADOS (2 veces cada uno) para el plugin "Shipment Tracking"
        query_insert = text("""
            INSERT INTO wpyz_postmeta (post_id, meta_key, meta_value)
            VALUES (:order_id, :key, :value)
        """)
        
        for key, value in [
            ('_tracking_number', tracking_number),
            ('_tracking_provider', shipping_provider),
            ('_wc_shipment_tracking_items', serialized_items)
        ]:
            db.session.execute(query_insert, {'order_id': order_id, 'key': key, 'value': value})
            db.session.execute(query_insert, {'order_id': order_id, 'key': key, 'value': value})

        # 3. GUARDAR EN HPOS (wpyz_wc_orders_meta)
        query_hpos = text("""
            INSERT INTO wpyz_wc_orders_meta (order_id, meta_key, meta_value)
            VALUES (:order_id, '_wc_shipment_tracking_items', :serialized_items)
            ON DUPLICATE KEY UPDATE meta_value = :serialized_items
        """)
        db.session.execute(query_hpos, {
            'order_id': order_id,
            'serialized_items': serialized_items
        })

        db.session.commit()

        # Obtener número de pedido
        order_number = order.get_meta('_order_number')
        if not order_number:
            order_number = f"#{order_id}"

        current_app.logger.info(
            f"Tracking agregado a pedido {order_number}: {tracking_number} ({shipping_provider}) "
            f"por {current_user.username}"
        )

        return jsonify({
            'success': True,
            'message': f'Tracking agregado exitosamente al pedido {order_number}',
            'order_id': order_id,
            'order_number': order_number,
            'tracking_number': tracking_number,
            'shipping_provider': shipping_provider,
            'date_shipped': date_shipped,
            'status_changed': mark_as_shipped
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error agregando tracking: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================
# TRACKING MASIVO CHAMO/DINSIDES
# ============================================


@bp.route('/api/bulk-tracking-simple', methods=['POST'])
@login_required
@master_required
def bulk_tracking_simple():
    """
    Procesa tracking masivo para CHAMO o DINSIDES con mensaje dinámico.

    A diferencia de Shalom que requiere un Excel con claves, CHAMO y DINSIDES
    usan mensajes de tracking con fecha seleccionable.

    Request JSON:
    {
        "orders": [41608, 41609, 41610],  # Lista de order IDs
        "column": "chamo" | "dinsides",
        "shipping_date": "2025-01-21"  # Fecha de envío seleccionada
    }

    Returns:
        JSON con resultado del proceso masivo
    """
    import time
    import phpserialize

    # Función para formatear fecha a texto legible (ej: "21 de enero")
    def format_date_spanish(date_str):
        months = [
            'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
            'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre'
        ]
        try:
            year, month, day = date_str.split('-')
            day_num = int(day)
            month_name = months[int(month) - 1]
            return f"{day_num} de {month_name}"
        except:
            return date_str

    try:
        data = request.get_json()
        order_ids = data.get('orders', [])
        column = data.get('column')
        shipping_date = data.get('shipping_date')

        if not order_ids:
            return jsonify({'success': False, 'error': 'No se especificaron pedidos'}), 400

        if column not in ['chamo', 'dinsides']:
            return jsonify({'success': False, 'error': 'Columna inválida'}), 400

        if not shipping_date:
            return jsonify({'success': False, 'error': 'No se especificó fecha de envío'}), 400

        # Formatear fecha para el mensaje
        fecha_formateada = format_date_spanish(shipping_date)

        # Plantillas de mensajes con fecha dinámica
        message_templates = {
            'chamo': f"Hola, somos izistore. Su pedido estará llegando el {fecha_formateada} entre las 11:00 am y 7:00 pm.",
            'dinsides': f"Hola, somos izistore. Su pedido está programado para ser entregado el {fecha_formateada} entre las 11:00 AM y 7:00 PM."
        }

        # Proveedores por columna
        providers = {
            'chamo': "Motorizado Izi",
            'dinsides': "Dinsides Courier"
        }

        tracking_message = message_templates[column]
        shipping_provider = providers[column]
        date_shipped = shipping_date  # Usar la fecha seleccionada
        timestamp = int(datetime.utcnow().timestamp())

        # Inicializar API de WooCommerce
        from woocommerce import API
        wc_api = API(
            url=current_app.config['WC_API_URL'],
            consumer_key=current_app.config['WC_CONSUMER_KEY'],
            consumer_secret=current_app.config['WC_CONSUMER_SECRET'],
            version="wc/v3",
            timeout=15
        )

        resultados = []
        exitosos = 0
        fallidos = 0

        current_app.logger.info(f"[BULK-TRACKING-SIMPLE] Iniciando proceso para {len(order_ids)} pedidos de {column.upper()}")

        for order_id in order_ids:
            try:
                # Validar que el pedido existe
                order = Order.query.get(order_id)
                if not order:
                    fallidos += 1
                    resultados.append({
                        'order_id': order_id,
                        'success': False,
                        'error': 'Pedido no encontrado'
                    })
                    continue

                # Obtener número de pedido
                order_number = order.get_meta('_order_number') or f"#{order_id}"

                # Crear tracking items en formato PHP serializado
                tracking_items = [{
                    'tracking_number': tracking_message,
                    'tracking_provider': shipping_provider,
                    'custom_tracking_provider': '',
                    'custom_tracking_link': '',
                    'date_shipped': date_shipped,
                    'tracking_id': str(timestamp)
                }]
                serialized_items = phpserialize.dumps(tracking_items).decode('utf-8')

                # 1. ACTUALIZAR VÍA API (cambia estado y envía email)
                try:
                    payment_method = order.get_meta('_payment_method') or order.payment_method
                    wc_api.put(f"orders/{order_id}", {
                        "status": "completed",
                        "payment_method": payment_method,
                        "meta_data": [
                            {"key": "_tracking_number", "value": tracking_message},
                            {"key": "_tracking_provider", "value": shipping_provider},
                            {"key": "_wc_shipment_tracking_items", "value": tracking_items}
                        ]
                    })
                    current_app.logger.info(f"[BULK-TRACKING-SIMPLE] Order {order_number}: API update success")
                    order.status = 'wc-completed'
                    order.date_updated_gmt = datetime.utcnow()

                except Exception as api_err:
                    current_app.logger.error(f"[BULK-TRACKING-SIMPLE] Order {order_number}: API Error - {str(api_err)}")
                    order.status = 'wc-completed'
                    order.date_updated_gmt = datetime.utcnow()

                # 2. GUARDAR EN BASE DE DATOS (duplicados para plugin)
                query_delete_old = text("""
                    DELETE FROM wpyz_postmeta
                    WHERE post_id = :order_id
                      AND meta_key IN ('_tracking_number', '_tracking_provider', '_wc_shipment_tracking_items')
                """)
                db.session.execute(query_delete_old, {'order_id': order_id})

                query_insert = text("""
                    INSERT INTO wpyz_postmeta (post_id, meta_key, meta_value)
                    VALUES (:order_id, :key, :value)
                """)

                for key, value in [
                    ('_tracking_number', tracking_message),
                    ('_tracking_provider', shipping_provider),
                    ('_wc_shipment_tracking_items', serialized_items)
                ]:
                    db.session.execute(query_insert, {'order_id': order_id, 'key': key, 'value': value})
                    db.session.execute(query_insert, {'order_id': order_id, 'key': key, 'value': value})

                # 3. GUARDAR EN HPOS
                query_hpos = text("""
                    INSERT INTO wpyz_wc_orders_meta (order_id, meta_key, meta_value)
                    VALUES (:order_id, '_wc_shipment_tracking_items', :serialized_items)
                    ON DUPLICATE KEY UPDATE meta_value = :serialized_items
                """)
                db.session.execute(query_hpos, {
                    'order_id': order_id,
                    'serialized_items': serialized_items
                })

                db.session.commit()

                exitosos += 1
                resultados.append({
                    'order_id': order_id,
                    'order_number': order_number,
                    'success': True
                })

                current_app.logger.info(
                    f"[BULK-TRACKING-SIMPLE] Tracking asignado a {order_number} ({column.upper()}) por {current_user.username}"
                )

                # Rate limiting - esperar 1 segundo entre pedidos
                if len(order_ids) > 1:
                    time.sleep(1)

            except Exception as e:
                db.session.rollback()
                fallidos += 1
                resultados.append({
                    'order_id': order_id,
                    'success': False,
                    'error': str(e)
                })
                current_app.logger.error(f"[BULK-TRACKING-SIMPLE] Error procesando order {order_id}: {str(e)}")

        current_app.logger.info(
            f"[BULK-TRACKING-SIMPLE] Proceso completado: {exitosos} exitosos, {fallidos} fallidos"
        )

        return jsonify({
            'success': True,
            'exitosos': exitosos,
            'fallidos': fallidos,
            'total': len(order_ids),
            'resultados': resultados
        })

    except Exception as e:
        current_app.logger.error(f"Error en bulk_tracking_simple: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# TRACKING MASIVO SHALOM
# ============================================


def validar_clave_shalom(clave, dni, telefono):
    """
    Valida y corrige la CLAVE de Shalom según las reglas de negocio.

    Reglas:
    1. Si los 4 dígitos son iguales (ej: 7777), usar los 2 primeros dígitos en lugar de los últimos
    2. No permitir patrones X0X0 (1010, 2020, 3030...9090)
    3. No permitir patrones X000 (1000, 2000, 3000...9000)

    Args:
        clave: CLAVE inicial (4 dígitos)
        dni: DNI completo del destinatario
        telefono: Teléfono completo del destinatario

    Returns:
        CLAVE validada/corregida
    """
    if len(clave) != 4:
        return clave

    # Patrones no permitidos
    patrones_x0x0 = ['1010', '2020', '3030', '4040', '5050', '6060', '7070', '8080', '9090']
    patrones_x000 = ['1000', '2000', '3000', '4000', '5000', '6000', '7000', '8000', '9000']
    patrones_otros = ['1234']

    # Combinar todos los patrones no permitidos
    patrones_invalidos = patrones_x0x0 + patrones_x000 + patrones_otros

    # Verificar si los 4 dígitos son iguales
    todos_iguales = len(set(clave)) == 1

    # Verificar si es un patrón no permitido
    es_patron_invalido = clave in patrones_invalidos

    # Si necesita corrección, usar los 2 primeros dígitos en lugar de los últimos
    if todos_iguales or es_patron_invalido:
        if len(dni) >= 2 and len(telefono) >= 2:
            clave_alternativa = dni[:2] + telefono[:2]
            # Verificar que la alternativa tampoco sea inválida
            if clave_alternativa not in patrones_invalidos and len(set(clave_alternativa)) > 1:
                return clave_alternativa

    return clave


@bp.route('/bulk-tracking')
@login_required
@master_required
def bulk_tracking_page():
    """
    Página para asignación masiva de tracking Shalom.
    """
    return render_template('dispatch_bulk_tracking.html')


@bp.route('/api/bulk-tracking/preview', methods=['POST'])
@login_required
@master_required
def bulk_tracking_preview():
    """
    Lee el archivo Excel de Shalom subido por el usuario y muestra preview de los envíos.
    Hace matching por DNI con pedidos Shalom en estado processing.

    Returns:
        JSON con lista de envíos y su estado de matching
    """
    try:
        from openpyxl import load_workbook
        from io import BytesIO

        # Verificar que se envió un archivo
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No se envió ningún archivo'
            }), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'No se seleccionó ningún archivo'
            }), 400

        # Validar extensión
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({
                'success': False,
                'error': 'El archivo debe ser un Excel (.xlsx o .xls)'
            }), 400

        excel_filename = file.filename
        current_app.logger.info(f"[BULK-TRACKING] Procesando archivo: {excel_filename}")

        # Cargar el Excel desde memoria (sin guardar en disco)
        workbook = load_workbook(BytesIO(file.read()), data_only=True)
        sheet = workbook.active

        # Parsear envíos (cada 26 filas)
        envios = []
        fila = 1
        envio_num = 1

        while True:
            # Verificar si hay datos en la primera celda del bloque
            celda_inicio = sheet[f'A{fila}'].value
            if not celda_inicio or not str(celda_inicio).strip():
                break

            # Extraer datos del bloque de 26 filas
            dni_destinatario = str(sheet[f'A{fila + 13}'].value or '').strip()  # A14
            tel_destinatario = str(sheet[f'A{fila + 14}'].value or '').strip()  # A15
            orden_shalom = str(sheet[f'A{fila + 22}'].value or '').strip()      # A23
            codigo_shalom = str(sheet[f'A{fila + 23}'].value or '').strip()     # A24

            # Construir la CLAVE (últimos 2 dígitos de DNI destinatario + últimos 2 de teléfono destinatario)
            clave = ''
            if len(dni_destinatario) >= 2 and len(tel_destinatario) >= 2:
                clave = dni_destinatario[-2:] + tel_destinatario[-2:]

                # Validar y corregir CLAVE si es necesario
                clave = validar_clave_shalom(clave, dni_destinatario, tel_destinatario)

            # Construir tracking number completo
            tracking_number = f"{orden_shalom} {codigo_shalom} CLAVE: {clave}"

            envio = {
                'envio_num': envio_num,
                'fila_inicio': fila,
                'dni_destinatario': dni_destinatario,
                'tel_destinatario': tel_destinatario,
                'orden_shalom': orden_shalom,
                'codigo_shalom': codigo_shalom,
                'clave': clave,
                'tracking_number': tracking_number,
                'pedido_id': None,
                'pedido_numero': None,
                'cliente_nombre': None,
                'estado_pedido': None,
                'validacion': 'pending',
                'mensaje': ''
            }

            envios.append(envio)
            fila += 26
            envio_num += 1

            # Límite de seguridad
            if envio_num > 100:
                break

        workbook.close()

        if not envios:
            return jsonify({
                'success': False,
                'error': 'No se encontraron envíos en el archivo Excel'
            }), 400

        # Obtener todos los pedidos Shalom en estado processing
        shalom_orders = get_shalom_orders_by_dni()

        # Hacer matching por DNI
        total_ok = 0
        total_ya_completado = 0
        total_no_encontrado = 0

        for envio in envios:
            dni = envio['dni_destinatario']

            if not dni or len(dni) < 8:
                envio['validacion'] = 'error'
                envio['mensaje'] = 'DNI inválido'
                total_no_encontrado += 1
                continue

            if dni in shalom_orders:
                pedido = shalom_orders[dni]
                envio['pedido_id'] = pedido['id']
                envio['pedido_numero'] = pedido['numero']
                envio['cliente_nombre'] = pedido['cliente']
                envio['estado_pedido'] = pedido['estado']

                if pedido['estado'] == 'wc-completed':
                    envio['validacion'] = 'warning'
                    envio['mensaje'] = 'Pedido ya completado'
                    total_ya_completado += 1
                else:
                    envio['validacion'] = 'ok'
                    envio['mensaje'] = 'Listo para procesar'
                    total_ok += 1
            else:
                envio['validacion'] = 'error'
                envio['mensaje'] = 'DNI no encontrado en pedidos Shalom'
                total_no_encontrado += 1

        return jsonify({
            'success': True,
            'envios': envios,
            'archivo': excel_filename,
            'resumen': {
                'total': len(envios),
                'ok': total_ok,
                'ya_completado': total_ya_completado,
                'no_encontrado': total_no_encontrado
            }
        })

    except Exception as e:
        current_app.logger.error(f"Error en preview de tracking masivo: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def get_shalom_orders_by_dni():
    """
    Obtiene todos los pedidos Shalom en estado processing,
    indexados por DNI del cliente.

    Returns:
        dict: {dni: {id, numero, cliente, estado}}
    """
    query = text("""
        SELECT
            o.id,
            COALESCE(om_number.meta_value, CONCAT('#', o.id)) as order_number,
            CONCAT(ba.first_name, ' ', ba.last_name) as customer_name,
            ba.company as customer_dni,
            o.status,
            (SELECT oi.order_item_name
             FROM wpyz_woocommerce_order_items oi
             WHERE oi.order_id = o.id
               AND oi.order_item_type = 'shipping'
             LIMIT 1) as shipping_method
        FROM wpyz_wc_orders o
        LEFT JOIN wpyz_wc_orders_meta om_number
            ON o.id = om_number.order_id
            AND om_number.meta_key = '_order_number'
        LEFT JOIN wpyz_wc_order_addresses ba
            ON o.id = ba.order_id
            AND ba.address_type = 'billing'
        WHERE o.status IN ('wc-processing', 'wc-completed')
          AND o.type = 'shop_order'
        ORDER BY o.date_created_gmt DESC
    """)

    results = db.session.execute(query).fetchall()

    # Indexar por DNI (solo pedidos Shalom)
    orders_by_dni = {}
    for row in results:
        shipping_method = row[5] or ''
        # Verificar si es pedido Shalom
        if 'shalom' in shipping_method.lower():
            dni = str(row[3] or '').strip()
            if dni and len(dni) >= 8:
                # Si hay múltiples pedidos con el mismo DNI, quedarse con el más reciente (processing primero)
                if dni not in orders_by_dni or row[4] == 'wc-processing':
                    orders_by_dni[dni] = {
                        'id': row[0],
                        'numero': row[1],
                        'cliente': row[2],
                        'estado': row[4]
                    }

    return orders_by_dni


@bp.route('/api/bulk-tracking/process', methods=['POST'])
@login_required
@master_required
def bulk_tracking_process():
    """
    Procesa los envíos seleccionados y asigna tracking a cada pedido.

    Request Body:
        {
            "envios": [
                {
                    "pedido_id": 41608,
                    "tracking_number": "N° de orden: 68529922 Código: JN79 CLAVE: 7014"
                },
                ...
            ],
            "fecha_envio": "2026-01-15"  // Opcional
        }

    Returns:
        JSON con resultados del procesamiento
    """
    try:
        data = request.get_json()
        envios = data.get('envios', [])
        fecha_envio = data.get('fecha_envio')

        if not envios:
            return jsonify({
                'success': False,
                'error': 'No hay envíos para procesar'
            }), 400

        # Fecha de envío (por defecto hoy)
        if not fecha_envio:
            from datetime import date
            fecha_envio = date.today().strftime('%Y-%m-%d')

        resultados = []
        exitosos = 0
        fallidos = 0

        for envio in envios:
            pedido_id = envio.get('pedido_id')
            tracking_number = envio.get('tracking_number')

            if not pedido_id or not tracking_number:
                resultados.append({
                    'pedido_id': pedido_id,
                    'success': False,
                    'error': 'Datos incompletos'
                })
                fallidos += 1
                continue

            # Procesar cada pedido usando la misma lógica que add_tracking
            resultado = process_single_tracking(
                order_id=pedido_id,
                tracking_number=tracking_number,
                shipping_provider='Shalom',
                date_shipped=fecha_envio,
                mark_as_shipped=True
            )

            resultados.append(resultado)

            if resultado['success']:
                exitosos += 1
            else:
                fallidos += 1

            # Rate limiting: esperar 1 segundo entre pedidos
            import time
            time.sleep(1)

        return jsonify({
            'success': True,
            'resultados': resultados,
            'resumen': {
                'total': len(envios),
                'exitosos': exitosos,
                'fallidos': fallidos
            }
        })

    except Exception as e:
        current_app.logger.error(f"Error procesando tracking masivo: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def process_single_tracking(order_id, tracking_number, shipping_provider, date_shipped, mark_as_shipped):
    """
    Procesa un solo tracking (misma lógica que add_tracking pero retorna dict).

    Returns:
        dict con resultado del procesamiento
    """
    try:
        # Validar que el pedido existe
        order = Order.query.get(order_id)
        if not order:
            return {
                'pedido_id': order_id,
                'success': False,
                'error': f'Pedido {order_id} no encontrado'
            }

        # Timestamp actual
        timestamp = int(datetime.utcnow().timestamp())

        # Crear el array de tracking items
        tracking_items = [{
            'tracking_number': tracking_number,
            'tracking_provider': shipping_provider,
            'custom_tracking_provider': '',
            'custom_tracking_link': '',
            'date_shipped': date_shipped,
            'tracking_id': str(timestamp)
        }]

        # Serializar a formato PHP
        import phpserialize
        serialized_items = phpserialize.dumps(tracking_items).decode('utf-8')

        current_app.logger.info(f"[BULK-TRACKING] Order {order_id}: Asignando tracking {tracking_number}")

        # 1. ACTUALIZAR VÍA API (para disparar emails)
        if mark_as_shipped:
            try:
                from woocommerce import API
                wc_api = API(
                    url=current_app.config['WC_API_URL'],
                    consumer_key=current_app.config['WC_CONSUMER_KEY'],
                    consumer_secret=current_app.config['WC_CONSUMER_SECRET'],
                    version="wc/v3",
                    timeout=15
                )

                payment_method = order.get_meta('_payment_method') or order.payment_method

                wc_api.put(f"orders/{order_id}", {
                    "status": "completed",
                    "payment_method": payment_method,
                    "meta_data": [
                        {"key": "_tracking_number", "value": tracking_number},
                        {"key": "_tracking_provider", "value": shipping_provider},
                        {"key": "_wc_shipment_tracking_items", "value": tracking_items}
                    ]
                })

                current_app.logger.info(f"[BULK-TRACKING] Order {order_id}: API update success")
                order.status = 'wc-completed'
                order.date_updated_gmt = datetime.utcnow()

            except Exception as api_err:
                current_app.logger.error(f"[BULK-TRACKING] API Error para {order_id}: {str(api_err)}")
                # Continuar con actualización local
                order.status = 'wc-completed'
                order.date_updated_gmt = datetime.utcnow()

        # 2. GUARDAR EN BASE DE DATOS (Legacy)
        query_delete_old = text("""
            DELETE FROM wpyz_postmeta
            WHERE post_id = :order_id
              AND meta_key IN ('_tracking_number', '_tracking_provider', '_wc_shipment_tracking_items')
        """)
        db.session.execute(query_delete_old, {'order_id': order_id})

        query_insert = text("""
            INSERT INTO wpyz_postmeta (post_id, meta_key, meta_value)
            VALUES (:order_id, :key, :value)
        """)

        for key, value in [
            ('_tracking_number', tracking_number),
            ('_tracking_provider', shipping_provider),
            ('_wc_shipment_tracking_items', serialized_items)
        ]:
            db.session.execute(query_insert, {'order_id': order_id, 'key': key, 'value': value})

        # 3. GUARDAR EN HPOS
        query_hpos = text("""
            INSERT INTO wpyz_wc_orders_meta (order_id, meta_key, meta_value)
            VALUES (:order_id, '_wc_shipment_tracking_items', :serialized_items)
            ON DUPLICATE KEY UPDATE meta_value = :serialized_items
        """)
        db.session.execute(query_hpos, {
            'order_id': order_id,
            'serialized_items': serialized_items
        })

        # 4. REGISTRAR EN DISPATCH HISTORY
        # Obtener el número de orden primero
        order_number = order.get_meta('_order_number') or f"#{order_id}"

        history = DispatchHistory(
            order_id=order_id,
            order_number=order_number,
            previous_shipping_method='SHALOM',
            new_shipping_method='SHALOM',
            changed_by=current_user.username,
            dispatch_note=f'[TRACKING MASIVO] {tracking_number}'
        )
        db.session.add(history)

        db.session.commit()

        current_app.logger.info(
            f"[BULK-TRACKING] Tracking asignado a {order_number}: {tracking_number} "
            f"por {current_user.username}"
        )

        return {
            'pedido_id': order_id,
            'pedido_numero': order_number,
            'tracking_number': tracking_number,
            'success': True,
            'mensaje': 'Tracking asignado correctamente'
        }

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[BULK-TRACKING] Error en pedido {order_id}: {str(e)}")
        return {
            'pedido_id': order_id,
            'success': False,
            'error': str(e)
        }


# ============================================
# TRACKING MASIVO OLVA
# ============================================

@bp.route('/bulk-tracking-olva')
@login_required
@master_required
def bulk_tracking_olva_page():
    """
    Página para asignación masiva de tracking OLVA.
    """
    return render_template('dispatch_bulk_tracking_olva.html')


@bp.route('/api/bulk-tracking-olva/preview', methods=['POST'])
@login_required
@master_required
def bulk_tracking_olva_preview():
    """
    Lee el archivo Excel de OLVA subido por el usuario y muestra preview de los envíos.
    Hace matching por nombre del destinatario con pedidos OLVA en estado processing.

    Formato del Excel:
    - Columna E (desde E8): Tracking number completo
    - Columna H (desde H8): Nombre del destinatario
    - Columna F (desde F8): Estado del envío

    Returns:
        JSON con lista de envíos y su estado de matching
    """
    try:
        from openpyxl import load_workbook
        from io import BytesIO

        # Verificar que se envió un archivo
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No se envió ningún archivo'
            }), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'No se seleccionó ningún archivo'
            }), 400

        # Validar extensión
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({
                'success': False,
                'error': 'El archivo debe ser un Excel (.xlsx o .xls)'
            }), 400

        excel_filename = file.filename
        current_app.logger.info(f"[BULK-TRACKING-OLVA] Procesando archivo: {excel_filename}")

        # Cargar el Excel desde memoria
        workbook = load_workbook(BytesIO(file.read()), data_only=True)
        sheet = workbook.active

        # Parsear envíos desde fila 8
        envios = []
        fila = 8

        while True:
            # Verificar si hay datos en la columna E (tracking)
            tracking_full = sheet[f'E{fila}'].value
            if not tracking_full or not str(tracking_full).strip():
                break

            # Extraer datos
            tracking_full = str(tracking_full).strip()
            nombre_destinatario = str(sheet[f'H{fila}'].value or '').strip()
            estado = str(sheet[f'F{fila}'].value or '').strip()

            # Obtener últimos 6 dígitos del tracking
            tracking_last_6 = tracking_full[-6:] if len(tracking_full) >= 6 else tracking_full

            # Determinar si lleva clave (si el estado contiene "Tienda" o "Agente")
            lleva_clave = estado_requiere_clave(estado)

            # Construir mensaje de tracking
            if lleva_clave:
                tracking_message = f"Nro. de tracking: {tracking_last_6}\nClave: 7747"
            else:
                tracking_message = f"Nro. de tracking: {tracking_last_6}"

            envio = {
                'fila': fila,
                'tracking_full': tracking_full,
                'tracking_last_6': tracking_last_6,
                'nombre_destinatario': nombre_destinatario,
                'estado': estado,
                'lleva_clave': lleva_clave,
                'tracking_message': tracking_message,
                'pedido_id': None,
                'pedido_numero': None,
                'cliente_nombre': None,
                'estado_pedido': None,
                'validacion': 'pending',
                'mensaje': ''
            }

            envios.append(envio)
            fila += 1

            # Límite de seguridad
            if len(envios) > 200:
                break

        workbook.close()

        if not envios:
            return jsonify({
                'success': False,
                'error': 'No se encontraron envíos en el archivo Excel (desde fila 8)'
            }), 400

        # Obtener todos los pedidos OLVA en estado processing
        olva_orders = get_olva_orders_by_name()

        # Hacer matching por nombre
        total_ok = 0
        total_ya_completado = 0
        total_no_encontrado = 0

        for envio in envios:
            nombre = envio['nombre_destinatario']

            if not nombre or len(nombre) < 3:
                envio['validacion'] = 'error'
                envio['mensaje'] = 'Nombre inválido'
                total_no_encontrado += 1
                continue

            # Buscar pedido por nombre (al menos un nombre + un apellido)
            pedido = match_pedido_por_nombre(nombre, olva_orders)

            if pedido:
                envio['pedido_id'] = pedido['id']
                envio['pedido_numero'] = pedido['numero']
                envio['cliente_nombre'] = pedido['cliente']
                envio['estado_pedido'] = pedido['estado']

                if pedido['estado'] == 'wc-completed':
                    envio['validacion'] = 'warning'
                    envio['mensaje'] = 'Pedido ya completado'
                    total_ya_completado += 1
                else:
                    envio['validacion'] = 'ok'
                    envio['mensaje'] = 'Listo para procesar'
                    total_ok += 1
            else:
                envio['validacion'] = 'error'
                envio['mensaje'] = 'Nombre no encontrado en pedidos OLVA'
                total_no_encontrado += 1

        return jsonify({
            'success': True,
            'envios': envios,
            'archivo': excel_filename,
            'resumen': {
                'total': len(envios),
                'ok': total_ok,
                'ya_completado': total_ya_completado,
                'no_encontrado': total_no_encontrado
            }
        })

    except Exception as e:
        current_app.logger.error(f"Error en preview de tracking masivo OLVA: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def estado_requiere_clave(estado):
    """
    Verifica si el estado del envío requiere clave (7747).

    Retorna True si el estado contiene "Tienda" o "Agente" (case insensitive).

    Args:
        estado: String con el estado del envío

    Returns:
        bool: True si requiere clave, False en caso contrario
    """
    if not estado:
        return False

    estado_lower = estado.lower()
    return 'tienda' in estado_lower or 'agente' in estado_lower


def get_olva_orders_by_name():
    """
    Obtiene todos los pedidos OLVA en estado processing,
    indexados por nombre del cliente (normalizado).

    Returns:
        dict: {nombre_normalizado: {id, numero, cliente, estado}}
    """
    query = text("""
        SELECT
            o.id,
            COALESCE(om_number.meta_value, CONCAT('#', o.id)) as order_number,
            CONCAT(ba.first_name, ' ', ba.last_name) as customer_name,
            o.status,
            (SELECT oi.order_item_name
             FROM wpyz_woocommerce_order_items oi
             WHERE oi.order_id = o.id
               AND oi.order_item_type = 'shipping'
             LIMIT 1) as shipping_method
        FROM wpyz_wc_orders o
        LEFT JOIN wpyz_wc_orders_meta om_number
            ON o.id = om_number.order_id
            AND om_number.meta_key = '_order_number'
        LEFT JOIN wpyz_wc_order_addresses ba
            ON o.id = ba.order_id
            AND ba.address_type = 'billing'
        WHERE o.status IN ('wc-processing', 'wc-completed')
          AND o.type = 'shop_order'
        ORDER BY o.date_created_gmt DESC
    """)

    results = db.session.execute(query).fetchall()

    # Indexar por nombre normalizado (solo pedidos OLVA)
    orders_by_name = {}
    for row in results:
        shipping_method = row[4] or ''
        # Verificar si es pedido OLVA
        if 'olva' in shipping_method.lower():
            nombre = str(row[2] or '').strip()
            if nombre and len(nombre) >= 3:
                nombre_normalizado = normalizar_nombre(nombre)
                # Si hay múltiples pedidos con el mismo nombre, quedarse con el más reciente (processing primero)
                if nombre_normalizado not in orders_by_name or row[3] == 'wc-processing':
                    orders_by_name[nombre_normalizado] = {
                        'id': row[0],
                        'numero': row[1],
                        'cliente': row[2],
                        'estado': row[3],
                        'nombre_original': nombre
                    }

    return orders_by_name


def normalizar_nombre(nombre):
    """
    Normaliza un nombre para facilitar comparaciones.
    - Convierte a minúsculas
    - Elimina acentos
    - Elimina espacios extras

    Args:
        nombre: String con el nombre a normalizar

    Returns:
        str: Nombre normalizado
    """
    if not nombre:
        return ''

    # Convertir a minúsculas
    nombre = nombre.lower()

    # Eliminar acentos
    nombre = ''.join(
        c for c in unicodedata.normalize('NFD', nombre)
        if unicodedata.category(c) != 'Mn'
    )

    # Eliminar espacios extras y normalizar
    nombre = ' '.join(nombre.split())

    return nombre


def match_pedido_por_nombre(nombre_excel, orders_by_name):
    """
    Busca un pedido por nombre, requiriendo al menos coincidencia de un nombre + un apellido.

    Args:
        nombre_excel: Nombre del destinatario desde el Excel
        orders_by_name: Diccionario de pedidos indexados por nombre normalizado

    Returns:
        dict: Pedido encontrado o None
    """
    if not nombre_excel or not orders_by_name:
        return None

    # Normalizar el nombre del Excel
    nombre_normalizado = normalizar_nombre(nombre_excel)

    # Separar en palabras (nombres y apellidos)
    palabras_excel = nombre_normalizado.split()

    # Si tiene menos de 2 palabras, no podemos hacer match (necesitamos al menos nombre + apellido)
    if len(palabras_excel) < 2:
        return None

    # Buscar en todos los pedidos
    mejores_matches = []

    for nombre_pedido_norm, pedido in orders_by_name.items():
        palabras_pedido = nombre_pedido_norm.split()

        # Contar cuántas palabras del Excel coinciden con el pedido
        coincidencias = 0
        for palabra_excel in palabras_excel:
            if palabra_excel in palabras_pedido:
                coincidencias += 1

        # Requerir al menos 2 palabras coincidentes (nombre + apellido)
        if coincidencias >= 2:
            mejores_matches.append({
                'pedido': pedido,
                'coincidencias': coincidencias,
                'score': coincidencias / max(len(palabras_excel), len(palabras_pedido))
            })

    # Si hay matches, devolver el mejor
    if mejores_matches:
        # Ordenar por score (mayor primero)
        mejores_matches.sort(key=lambda x: x['score'], reverse=True)
        return mejores_matches[0]['pedido']

    return None


@bp.route('/api/bulk-tracking-olva/process', methods=['POST'])
@login_required
@master_required
def bulk_tracking_olva_process():
    """
    Procesa los envíos OLVA seleccionados y asigna tracking a cada pedido.

    Request Body:
        {
            "envios": [
                {
                    "pedido_id": 41608,
                    "tracking_number": "Nro. de tracking: 139656\nClave: 7747"
                },
                ...
            ],
            "fecha_envio": "2026-01-15"  // Opcional
        }

    Returns:
        JSON con resultados del procesamiento
    """
    try:
        data = request.get_json()
        envios = data.get('envios', [])
        fecha_envio = data.get('fecha_envio')

        if not envios:
            return jsonify({
                'success': False,
                'error': 'No hay envíos para procesar'
            }), 400

        # Fecha de envío (por defecto hoy)
        if not fecha_envio:
            from datetime import date
            fecha_envio = date.today().strftime('%Y-%m-%d')

        resultados = []
        exitosos = 0
        fallidos = 0

        for envio in envios:
            pedido_id = envio.get('pedido_id')
            tracking_number = envio.get('tracking_number')

            if not pedido_id or not tracking_number:
                resultados.append({
                    'pedido_id': pedido_id,
                    'success': False,
                    'error': 'Datos incompletos'
                })
                fallidos += 1
                continue

            # Procesar cada pedido usando la misma lógica que add_tracking
            resultado = process_single_tracking(
                order_id=pedido_id,
                tracking_number=tracking_number,
                shipping_provider='Olva Courier',
                date_shipped=fecha_envio,
                mark_as_shipped=True
            )

            resultados.append(resultado)

            if resultado['success']:
                exitosos += 1
            else:
                fallidos += 1

            # Rate limiting: esperar 1 segundo entre pedidos
            import time
            time.sleep(1)

        return jsonify({
            'success': True,
            'resultados': resultados,
            'resumen': {
                'total': len(envios),
                'exitosos': exitosos,
                'fallidos': fallidos
            }
        })

    except Exception as e:
        current_app.logger.error(f"Error procesando tracking masivo OLVA: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
