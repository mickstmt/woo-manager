# app/routes/reports.py
from flask import Blueprint, render_template, request, jsonify, current_app, send_file
from flask_login import login_required, current_user
from app.routes.auth import admin_required
from app import db
from app.models import TipoCambio
from datetime import datetime, timedelta, date
from decimal import Decimal
from sqlalchemy import text, func
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

bp = Blueprint('reports', __name__, url_prefix='/reports')


def normalize_payment_method(payment_method_value, is_whatsapp=False):
    """
    Transforma el payment_method en plataforma con formato:
    - WEBSITE IZI EFECTIVO / WEBSITE IZI TARJETA (pedidos WooCommerce)
    - WHATSAPP IZI EFECTIVO / WHATSAPP IZI TARJETA (pedidos WhatsApp)

    Args:
        payment_method_value: Valor de payment_method o payment_method_title
        is_whatsapp: True si el pedido es de WhatsApp (tiene _order_number W-XXXXX)

    Returns:
        str: Plataforma normalizada
    """
    if not payment_method_value or payment_method_value == 'N/A':
        return 'N/A'

    value = str(payment_method_value).lower()

    # Determinar prefijo según origen del pedido
    prefix = 'WHATSAPP IZI' if is_whatsapp else 'WEBSITE IZI'

    # Clasificar como TARJETA o EFECTIVO
    # TARJETA: tarjetas de débito/crédito y Mercado Pago
    tarjeta_keywords = [
        'tarjeta',
        'card',
        'débito',
        'debito',
        'crédito',
        'credito',
        'mercado pago',
        'mercadopago',
        'mercado_pago',
        'woo-mercado-pago',
        'cuotas'
    ]

    # EFECTIVO: QR, Yape, Plin, Banca Virtual, Transferencias, Efectivo
    efectivo_keywords = [
        'qr',
        'yape',
        'plin',
        'bacs',
        'banca virtual',
        'transferencia',
        'efectivo',
        'cod',
        'cash',
        'cheque'
    ]

    # Verificar si es TARJETA
    for keyword in tarjeta_keywords:
        if keyword in value:
            return f'{prefix} TARJETA'

    # Verificar si es EFECTIVO
    for keyword in efectivo_keywords:
        if keyword in value:
            return f'{prefix} EFECTIVO'

    # Si no coincide con ninguna categoría, retornar N/A
    return 'N/A'


@bp.route('/')
@login_required
def index():
    """
    Página principal de reportes con dashboard
    """
    return render_template('reports.html', title='Reportes')


@bp.route('/api/summary')
@login_required
def api_summary():
    """
    Obtener resumen general de métricas clave

    Query params:
    - start_date: fecha inicio (YYYY-MM-DD)
    - end_date: fecha fin (YYYY-MM-DD)
    """
    try:
        # Obtener fechas de filtro
        start_date = request.args.get('start_date', type=str)
        end_date = request.args.get('end_date', type=str)

        # Si no se especifican, usar hoy
        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        if not start_date:
            start_date = datetime.now().strftime('%Y-%m-%d')

        # Query para métricas generales
        # IMPORTANTE: Convertir UTC a hora Perú (UTC-5) antes de comparar fechas
        query = text("""
            SELECT
                COUNT(DISTINCT o.id) as total_orders,
                COALESCE(SUM(o.total_amount), 0) as total_sales,
                COALESCE(AVG(o.total_amount), 0) as avg_order_value,
                COUNT(DISTINCT CASE WHEN o.status = 'wc-completed' THEN o.id END) as completed_orders,
                COUNT(DISTINCT CASE WHEN o.status = 'wc-cancelled' THEN o.id END) as cancelled_orders,
                COUNT(DISTINCT CASE WHEN o.status = 'wc-processing' THEN o.id END) as processing_orders
            FROM wpyz_wc_orders o
            INNER JOIN wpyz_wc_orders_meta om ON o.id = om.order_id AND om.meta_key = '_order_number'
            WHERE DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR)) BETWEEN :start_date AND :end_date
                AND o.status != 'trash'
        """)

        result = db.session.execute(query, {
            'start_date': start_date,
            'end_date': end_date
        }).fetchone()

        # Query para total de descuentos
        discount_query = text("""
            SELECT COALESCE(SUM(CAST(meta_value AS DECIMAL(10,2))), 0) as total_discounts
            FROM wpyz_wc_orders_meta om
            INNER JOIN wpyz_wc_orders o ON om.order_id = o.id
            WHERE om.meta_key = '_wc_discount_amount'
                AND DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR)) BETWEEN :start_date AND :end_date
                AND o.status != 'trash'
        """)

        discount_result = db.session.execute(discount_query, {
            'start_date': start_date,
            'end_date': end_date
        }).fetchone()

        return jsonify({
            'success': True,
            'data': {
                'total_orders': result[0] or 0,
                'total_sales': float(result[1] or 0),
                'avg_order_value': float(result[2] or 0),
                'completed_orders': result[3] or 0,
                'cancelled_orders': result[4] or 0,
                'processing_orders': result[5] or 0,
                'total_discounts': float(discount_result[0] or 0),
                'period': {
                    'start': start_date,
                    'end': end_date
                }
            }
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/api/sales-by-day')
@login_required
def api_sales_by_day():
    """
    Ventas agrupadas por día para gráfico de tendencia
    """
    try:
        start_date = request.args.get('start_date', type=str)
        end_date = request.args.get('end_date', type=str)

        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        if not start_date:
            start_date = datetime.now().strftime('%Y-%m-%d')

        query = text("""
            SELECT
                DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR)) as date,
                COUNT(o.id) as orders,
                COALESCE(SUM(o.total_amount), 0) as total
            FROM wpyz_wc_orders o
            INNER JOIN wpyz_wc_orders_meta om ON o.id = om.order_id AND om.meta_key = '_order_number'
            WHERE DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR)) BETWEEN :start_date AND :end_date
                AND o.status != 'trash'
            GROUP BY DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR))
            ORDER BY date ASC
        """)

        results = db.session.execute(query, {
            'start_date': start_date,
            'end_date': end_date
        }).fetchall()

        data = [{
            'date': str(row[0]),
            'orders': row[1],
            'total': float(row[2])
        } for row in results]

        return jsonify({
            'success': True,
            'data': data
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/api/top-products')
@login_required
def api_top_products():
    """
    Productos más vendidos en el período
    """
    try:
        start_date = request.args.get('start_date', type=str)
        end_date = request.args.get('end_date', type=str)
        limit = request.args.get('limit', 10, type=int)

        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        if not start_date:
            start_date = datetime.now().strftime('%Y-%m-%d')

        query = text("""
            SELECT
                p.post_title as product_name,
                SUM(CAST(oim.meta_value AS SIGNED)) as quantity_sold,
                COUNT(DISTINCT oi.order_id) as times_ordered
            FROM wpyz_woocommerce_order_items oi
            INNER JOIN wpyz_wc_orders o ON oi.order_id = o.id
            INNER JOIN wpyz_woocommerce_order_itemmeta oim ON oi.order_item_id = oim.order_item_id AND oim.meta_key = '_qty'
            INNER JOIN wpyz_woocommerce_order_itemmeta oim_product ON oi.order_item_id = oim_product.order_item_id AND oim_product.meta_key = '_product_id'
            INNER JOIN wpyz_posts p ON CAST(oim_product.meta_value AS SIGNED) = p.ID
            WHERE oi.order_item_type = 'line_item'
                AND DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR)) BETWEEN :start_date AND :end_date
                AND o.status != 'trash'
            GROUP BY p.ID, p.post_title
            ORDER BY quantity_sold DESC
            LIMIT :limit
        """)

        results = db.session.execute(query, {
            'start_date': start_date,
            'end_date': end_date,
            'limit': limit
        }).fetchall()

        data = [{
            'product_name': row[0],
            'quantity_sold': row[1],
            'times_ordered': row[2]
        } for row in results]

        return jsonify({
            'success': True,
            'data': data
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/api/sales-by-user')
@login_required
def api_sales_by_user():
    """
    Ventas por usuario creador
    """
    try:
        start_date = request.args.get('start_date', type=str)
        end_date = request.args.get('end_date', type=str)

        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        if not start_date:
            start_date = datetime.now().strftime('%Y-%m-%d')

        query = text("""
            SELECT
                om_created.meta_value as username,
                COUNT(o.id) as total_orders,
                COALESCE(SUM(o.total_amount), 0) as total_sales
            FROM wpyz_wc_orders o
            INNER JOIN wpyz_wc_orders_meta om_number ON o.id = om_number.order_id AND om_number.meta_key = '_order_number'
            LEFT JOIN wpyz_wc_orders_meta om_created ON o.id = om_created.order_id AND om_created.meta_key = '_created_by'
            WHERE DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR)) BETWEEN :start_date AND :end_date
                AND o.status != 'trash'
            GROUP BY om_created.meta_value
            ORDER BY total_sales DESC
        """)

        results = db.session.execute(query, {
            'start_date': start_date,
            'end_date': end_date
        }).fetchall()

        # Obtener nombres completos de usuarios
        from app.models import User
        data = []
        for row in results:
            username = row[0] or 'Desconocido'
            user = User.query.filter_by(username=username).first() if username != 'Desconocido' else None

            data.append({
                'username': username,
                'full_name': user.full_name if user and user.full_name else username,
                'total_orders': row[1],
                'total_sales': float(row[2])
            })

        return jsonify({
            'success': True,
            'data': data
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/api/status-distribution')
@login_required
def api_status_distribution():
    """
    Distribución de pedidos por estado
    """
    try:
        start_date = request.args.get('start_date', type=str)
        end_date = request.args.get('end_date', type=str)

        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        if not start_date:
            start_date = datetime.now().strftime('%Y-%m-%d')

        query = text("""
            SELECT
                o.status,
                COUNT(o.id) as count
            FROM wpyz_wc_orders o
            INNER JOIN wpyz_wc_orders_meta om ON o.id = om.order_id AND om.meta_key = '_order_number'
            WHERE DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR)) BETWEEN :start_date AND :end_date
                AND o.status != 'trash'
            GROUP BY o.status
            ORDER BY count DESC
        """)

        results = db.session.execute(query, {
            'start_date': start_date,
            'end_date': end_date
        }).fetchall()

        data = [{
            'status': row[0],
            'count': row[1]
        } for row in results]

        return jsonify({
            'success': True,
            'data': data
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/profits')
@login_required
def profits_report():
    """
    Página de reporte de ganancias
    """
    return render_template('reports_profits.html', title='Reporte de Ganancias')


@bp.route('/api/profits')
@login_required
def api_profits():
    """
    Versión OPTIMIZADA del reporte de ganancias.
    Elimina el problema N+1 y subconsultas redundantes.
    """
    try:
        start_date = request.args.get('start_date', type=str)
        end_date = request.args.get('end_date', type=str)
        source = request.args.get('source', type=str, default='all')

        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')

        # 1. Obtener todos los costos de FC en un diccionario para búsqueda rápida
        # Se asume que el SKU en fccost es de 7 caracteres y debe coincidir parcialmente con el SKU de Woo
        fc_costs_rows = db.session.execute(text("SELECT sku, FCLastCost FROM woo_products_fccost WHERE LENGTH(sku) = 7")).fetchall()
        fc_costs_map = {row[0]: float(row[1] or 0) for row in fc_costs_rows}

        # 2. Obtener tipos de cambio históricos necesarios
        tc_rows = db.session.execute(text("SELECT fecha, tasa_promedio FROM woo_tipo_cambio WHERE activo = TRUE ORDER BY fecha DESC")).fetchall()
        # Función auxiliar para obtener TC más cercano a una fecha
        def get_tc_for_date(order_date):
            order_date_str = str(order_date)
            for tc_date, rate in tc_rows:
                if str(tc_date) <= order_date_str:
                    return float(rate)
            return 1.0 # Fallback

        # 3. Construir filtros
        source_filter = ""
        if source == 'whatsapp':
            source_filter = "AND (om_numero.meta_value IS NOT NULL OR oext.id IS NOT NULL)"
        elif source == 'woocommerce':
            source_filter = "AND om_numero.meta_value IS NULL AND oext.id IS NULL"

        # 4. Consulta de Órdenes (Solo datos básicos)
        orders_sql = text(f"""
            SELECT
                o.id,
                COALESCE(om_numero.meta_value, CAST(o.id AS CHAR)) as numero_pedido,
                DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR)) as fecha,
                o.status,
                o.total_amount,
                ba.first_name,
                ba.last_name,
                COALESCE((
                    SELECT SUM(CAST(oim_shipping.meta_value AS DECIMAL(10,2)))
                    FROM wpyz_woocommerce_order_items oi_shipping
                    INNER JOIN wpyz_woocommerce_order_itemmeta oim_shipping ON oi_shipping.order_item_id = oim_shipping.order_item_id
                    WHERE oi_shipping.order_id = o.id
                        AND oi_shipping.order_item_type = 'shipping'
                        AND oim_shipping.meta_key = 'cost'
                ), 0) as costo_envio
            FROM wpyz_wc_orders o
            LEFT JOIN wpyz_wc_orders_meta om_numero ON o.id = om_numero.order_id AND om_numero.meta_key = '_order_number'
            LEFT JOIN wpyz_wc_order_addresses ba ON o.id = ba.order_id AND ba.address_type = 'billing'
            LEFT JOIN woo_orders_ext oext ON om_numero.meta_value COLLATE utf8mb4_unicode_ci = oext.order_number
            WHERE DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR)) BETWEEN :start_date AND :end_date
                AND o.status != 'trash'
                AND o.status NOT IN ('wc-cancelled', 'wc-refunded', 'wc-failed')
                {source_filter}
            ORDER BY fecha DESC, o.id DESC
        """)
        
        orders_results = db.session.execute(orders_sql, {'start_date': start_date, 'end_date': end_date}).fetchall()
        order_ids = [r[0] for r in orders_results]

        if not order_ids:
            return jsonify({'success': True, 'data': {'orders': [], 'summary': {}, 'period': {'start': start_date, 'end': end_date}}})

        # 5. Consulta masiva de items para todas las órdenes encontradas
        items_sql = text("""
            SELECT
                oi.order_id,
                oi.order_item_name,
                oim_qty.meta_value as qty,
                pm_sku.meta_value as sku
            FROM wpyz_woocommerce_order_items oi
            INNER JOIN wpyz_woocommerce_order_itemmeta oim_pid ON oi.order_item_id = oim_pid.order_item_id AND oim_pid.meta_key = '_product_id'
            INNER JOIN wpyz_woocommerce_order_itemmeta oim_qty ON oi.order_item_id = oim_qty.order_item_id AND oim_qty.meta_key = '_qty'
            LEFT JOIN wpyz_woocommerce_order_itemmeta oim_vid ON oi.order_item_id = oim_vid.order_item_id AND oim_vid.meta_key = '_variation_id'
            LEFT JOIN wpyz_postmeta pm_sku ON CAST(COALESCE(NULLIF(oim_vid.meta_value, '0'), oim_pid.meta_value) AS UNSIGNED) = pm_sku.post_id AND pm_sku.meta_key = '_sku'
            WHERE oi.order_id IN :order_ids AND oi.order_item_type = 'line_item'
        """)
        
        # SQLAlchemy bindparams for IN clause can be tricky with raw text, better to use formatting for small lists or bind with list
        items_results = db.session.execute(items_sql, {'order_ids': order_ids}).fetchall()
        
        # Agrupar items por pedido
        items_by_order = {}
        for row in items_results:
            oid, name, qty, sku = row
            if oid not in items_by_order: items_by_order[oid] = []
            
            # Calcular costo unitario buscando en el mapa (Lógica LIKE %sku% en Python)
            cost_unit = 0.0
            if sku:
                # Buscar si alguna llave del fc_costs_map (que son de longitud 7) está contenida en el SKU de Woo
                for fc_sku, cost in fc_costs_map.items():
                    if fc_sku in sku:
                        cost_unit = cost
                        break

            items_by_order[oid].append({
                'producto': name,
                'cantidad': int(qty or 0),
                'sku': sku or 'Sin SKU',
                'costo_unitario_usd': cost_unit,
                'costo_total_usd': cost_unit * int(qty or 0),
                'tiene_sku': bool(sku),
                'tiene_costo': cost_unit > 0
            })

        # 6. Combinar y calcular Totales en una sola pasada
        final_orders = []
        total_p = 0
        total_ventas_pen = 0.0
        total_costos_usd = 0.0
        total_costos_pen = 0.0
        total_envio_pen = 0.0
        total_ganancia_pen = 0.0

        for r in orders_results:
            oid, num, fecha, status, total_venta, fname, lname, c_envio = r
            tc = get_tc_for_date(fecha)
            items = items_by_order.get(oid, [])
            
            costo_pedido_usd = sum(i['costo_total_usd'] for i in items)
            costo_pedido_pen = costo_pedido_usd * tc
            c_envio = float(c_envio or 0)
            ganancia = float(total_venta or 0) - costo_pedido_pen - c_envio
            margen = (ganancia / float(total_venta or 1) * 100) if total_venta and float(total_venta) > 0 else 0

            # Solo incluir pedidos con items con costo (para coincidir con el HAVING anterior)
            # O si prefieres incluir todos, quitar este check. 
            # El original decía: HAVING costo_total_usd IS NOT NULL
            if costo_pedido_usd == 0 and not items: continue 

            final_orders.append({
                'pedido_id': oid,
                'numero_pedido': num or str(oid),
                'fecha_pedido': str(fecha),
                'estado': status,
                'total_venta_pen': float(total_venta or 0),
                'tipo_cambio': tc,
                'costo_total_usd': costo_pedido_usd,
                'costo_total_pen': costo_pedido_pen,
                'costo_envio_pen': c_envio,
                'ganancia_pen': ganancia,
                'margen_porcentaje': round(margen, 2),
                'cliente_nombre': fname or 'Sin nombre',
                'cliente_apellido': lname or '',
                'items': items,
                'total_items': len(items)
            })

            # Acumular resumen
            total_p += 1
            total_ventas_pen += float(total_venta or 0)
            total_costos_usd += costo_pedido_usd
            total_costos_pen += costo_pedido_pen
            total_envio_pen += c_envio
            total_ganancia_pen += ganancia

        return jsonify({
            'success': True,
            'data': {
                'orders': final_orders,
                'summary': {
                    'total_pedidos': total_p,
                    'ventas_totales_pen': round(total_ventas_pen, 2),
                    'costos_totales_usd': round(total_costos_usd, 2),
                    'costos_totales_pen': round(total_costos_pen, 2),
                    'costos_envio_totales_pen': round(total_envio_pen, 2),
                    'ganancias_totales_pen': round(total_ganancia_pen, 2),
                    'margen_promedio_porcentaje': round((total_ganancia_pen / total_ventas_pen * 100) if total_ventas_pen > 0 else 0, 2)
                },
                'period': {'start': start_date, 'end': end_date}
            }
        })

    except Exception as e:
        import traceback
        current_app.logger.error(f"Error en reporte ganancias optimizado: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/api/profits/externos')
@login_required
def api_profits_externos():
    """
    Reporte de ganancias para pedidos externos (woo_orders_ext)

    Query params:
    - start_date: fecha inicio (YYYY-MM-DD)
    - end_date: fecha fin (YYYY-MM-DD)

    Retorna mismo formato que api_profits para consistencia
    """
    try:
        start_date = request.args.get('start_date', type=str)
        end_date = request.args.get('end_date', type=str)

        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')

        # 1. Obtener todos los costos de FC en un diccionario para búsqueda rápida
        fc_costs_rows = db.session.execute(text("SELECT sku, FCLastCost FROM woo_products_fccost WHERE LENGTH(sku) = 7")).fetchall()
        fc_costs_map = {row[0]: float(row[1] or 0) for row in fc_costs_rows}

        # 2. Obtener tipos de cambio históricos necesarios
        tc_rows = db.session.execute(text("SELECT fecha, tasa_promedio FROM woo_tipo_cambio WHERE activo = TRUE ORDER BY fecha DESC")).fetchall()
        def get_tc_for_date(order_date):
            order_date_str = str(order_date)
            for tc_date, rate in tc_rows:
                if str(tc_date) <= order_date_str:
                    return float(rate)
            return 3.8 # Fallback

        # 3. Consulta de Órdenes (Solo datos básicos)
        orders_query = text("""
            SELECT
                oext.id as pedido_id,
                oext.order_number as numero_pedido,
                DATE(DATE_SUB(oext.date_created_gmt, INTERVAL 5 HOUR)) as fecha_pedido,
                oext.status as estado,
                COALESCE(oext.payment_method_title, oext.payment_method, 'N/A') as metodo_pago,
                oext.total_amount as total_venta_pen,
                COALESCE(oext.shipping_cost, 0) as costo_envio_pen,
                oext.customer_first_name as cliente_nombre,
                oext.customer_last_name as cliente_apellido
            FROM woo_orders_ext oext
            WHERE DATE(DATE_SUB(oext.date_created_gmt, INTERVAL 5 HOUR)) BETWEEN :start_date AND :end_date
                AND oext.status != 'trash'
                AND oext.status NOT IN ('wc-cancelled', 'wc-refunded', 'wc-failed')
            ORDER BY fecha_pedido DESC, oext.id DESC
            LIMIT 1000
        """)

        orders_results = db.session.execute(orders_query, {'start_date': start_date, 'end_date': end_date}).fetchall()
        order_ids = [r[0] for r in orders_results]

        if not order_ids:
            return jsonify({
                'success': True,
                'orders': [],
                'summary': {
                    'total_pedidos': 0,
                    'ventas_totales_pen': 0,
                    'costos_totales_usd': 0,
                    'costos_totales_pen': 0,
                    'costos_envio_totales_pen': 0,
                    'ganancias_totales_pen': 0,
                    'margen_promedio_porcentaje': 0
                }
            })

        # 4. Consulta masiva de items para todas las órdenes encontradas
        items_sql = text("""
            SELECT order_ext_id, product_name, product_sku, quantity
            FROM woo_orders_ext_items
            WHERE order_ext_id IN :order_ids
        """)
        items_results = db.session.execute(items_sql, {'order_ids': order_ids}).fetchall()
        
        # Agrupar items por pedido
        items_by_order = {}
        for row in items_results:
            oid, name, sku, qty = row
            if oid not in items_by_order: items_by_order[oid] = []
            
            # Buscar costo del SKU
            costo_unitario = 0.0
            tiene_sku = bool(sku)
            tiene_costo = False
            
            if sku:
                for fc_sku, cost in fc_costs_map.items():
                    if fc_sku in sku:
                        costo_unitario = cost
                        tiene_costo = True
                        break

            items_by_order[oid].append({
                'producto': name or 'Sin nombre',
                'sku': sku or 'N/A',
                'cantidad': int(qty or 0),
                'costo_unitario_usd': round(costo_unitario, 2),
                'costo_total_usd': round(costo_unitario * int(qty or 0), 2),
                'tiene_sku': tiene_sku,
                'tiene_costo': tiene_costo
            })

        # 5. Combinar y calcular
        orders_list = []
        total_ventas = 0
        total_costos_usd = 0
        total_costos_pen = 0
        total_envios = 0
        total_ganancias = 0

        for row in orders_results:
            pedido_id, numero_pedido, fecha_pedido, estado, metodo_pago, total_venta_pen, costo_envio_pen, c_nombre, c_apellido = row
            
            total_venta_pen = float(total_venta_pen or 0)
            costo_envio_pen = float(costo_envio_pen or 0)
            tipo_cambio = get_tc_for_date(fecha_pedido)
            
            items = items_by_order.get(pedido_id, [])
            costo_total_usd = sum(i['costo_total_usd'] for i in items)
            costo_total_pen = costo_total_usd * tipo_cambio
            
            ganancia_pen = total_venta_pen - costo_total_pen - costo_envio_pen
            margen_porcentaje = (ganancia_pen / total_venta_pen * 100) if total_venta_pen > 0 else 0

            orders_list.append({
                'pedido_id': pedido_id,
                'numero_pedido': numero_pedido or str(pedido_id),
                'fecha_pedido': fecha_pedido.isoformat() if fecha_pedido else None,
                'estado': estado,
                'metodo_pago': metodo_pago,
                'total_venta_pen': round(total_venta_pen, 2),
                'tipo_cambio': round(tipo_cambio, 2),
                'costo_total_usd': round(costo_total_usd, 2),
                'costo_total_pen': round(costo_total_pen, 2),
                'costo_envio_pen': round(costo_envio_pen, 2),
                'ganancia_pen': round(ganancia_pen, 2),
                'margen_porcentaje': round(margen_porcentaje, 2),
                'cliente_nombre': c_nombre or 'Sin nombre',
                'cliente_apellido': c_apellido or '',
                'total_items': len(items),
                'items': items
            })

            total_ventas += total_venta_pen
            total_costos_usd += costo_total_usd
            total_costos_pen += costo_total_pen
            total_envios += costo_envio_pen
            total_ganancias += ganancia_pen

        margen_promedio = (total_ganancias / total_ventas * 100) if total_ventas > 0 else 0

        return jsonify({
            'success': True,
            'orders': orders_list,
            'summary': {
                'total_pedidos': len(orders_list),
                'ventas_totales_pen': round(total_ventas, 2),
                'costos_totales_usd': round(total_costos_usd, 2),
                'costos_totales_pen': round(total_costos_pen, 2),
                'costos_envio_totales_pen': round(total_envios, 2),
                'ganancias_totales_pen': round(total_ganancias, 2),
                'margen_promedio_porcentaje': round(margen_promedio, 2)
            }
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/api/profits/externos/export')
@login_required
def export_profits_externos_excel():
    """
    Exportar reporte de ganancias de pedidos externos a Excel

    Query params:
    - start_date: fecha inicio (YYYY-MM-DD)
    - end_date: fecha fin (YYYY-MM-DD)
    - statuses: estados a incluir separados por coma (ej: 'wc-completed,wc-processing')
    """
    try:
        start_date = request.args.get('start_date', type=str)
        end_date = request.args.get('end_date', type=str)
        statuses_param = request.args.get('statuses', type=str, default='')

        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')

        # Procesar filtro de estados
        if statuses_param:
            status_list = [s.strip() for s in statuses_param.split(',') if s.strip()]
            if status_list:
                status_placeholders = ', '.join([f"'{s}'" for s in status_list])
                status_filter = f"AND oext.status IN ({status_placeholders})"
            else:
                status_filter = "AND oext.status NOT IN ('wc-cancelled', 'wc-refunded', 'wc-failed')"
        else:
            status_filter = "AND oext.status NOT IN ('wc-cancelled', 'wc-refunded', 'wc-failed')"

        # 1. Obtener todos los costos de FC en un diccionario para búsqueda rápida
        fc_costs_rows = db.session.execute(text("SELECT sku, FCLastCost FROM woo_products_fccost WHERE LENGTH(sku) = 7")).fetchall()
        fc_costs_map = {row[0]: float(row[1] or 0) for row in fc_costs_rows}

        # 2. Obtener tipos de cambio históricos necesarios
        tc_rows = db.session.execute(text("SELECT fecha, tasa_promedio FROM woo_tipo_cambio WHERE activo = TRUE ORDER BY fecha DESC")).fetchall()
        def get_tc_for_date(order_date):
            order_date_str = str(order_date)
            for tc_date, rate in tc_rows:
                if str(tc_date) <= order_date_str:
                    return float(rate)
            return 3.8 # Fallback

        # 3. Consulta de Órdenes (Solo datos básicos)
        orders_query_template = """
            SELECT
                oext.id as pedido_id,
                oext.order_number as numero_pedido,
                DATE(DATE_SUB(oext.date_created_gmt, INTERVAL 5 HOUR)) as fecha_pedido,
                oext.status as estado,
                COALESCE(oext.payment_method_title, oext.payment_method, 'N/A') as metodo_pago,
                oext.total_amount as total_venta_pen,
                COALESCE(oext.shipping_cost, 0) as costo_envio_pen,
                oext.customer_first_name as cliente_nombre,
                oext.customer_last_name as cliente_apellido
            FROM woo_orders_ext oext
            WHERE DATE(DATE_SUB(oext.date_created_gmt, INTERVAL 5 HOUR)) BETWEEN :start_date AND :end_date
                AND oext.status != 'trash'
                {status_filter}
            ORDER BY fecha_pedido DESC, oext.id DESC
            LIMIT 1000
        """

        orders_query = text(orders_query_template.replace('{status_filter}', status_filter))
        orders_results = db.session.execute(orders_query, {'start_date': start_date, 'end_date': end_date}).fetchall()
        order_ids = [r[0] for r in orders_results]

        if not order_ids:
            # Si no hay pedidos, devolver Excel vacío o manejar correspondientemente
            wb = Workbook()
            ws = wb.active
            ws.title = "Pedidos Externos"
            ws['A1'] = "No hay datos para el período seleccionado"
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f"ganancias_externos_{start_date}_{end_date}.xlsx")

        # 4. Consulta masiva de items para todas las órdenes encontradas
        items_sql = text("""
            SELECT order_ext_id, product_sku, quantity
            FROM woo_orders_ext_items
            WHERE order_ext_id IN :order_ids
        """)
        items_results = db.session.execute(items_sql, {'order_ids': order_ids}).fetchall()
        
        # Agrupar items por pedido (solo necesitamos el costo total por pedido para el Excel resumen)
        costo_total_usd_by_order = {}
        for row in items_results:
            oid, sku, qty = row
            if oid not in costo_total_usd_by_order: costo_total_usd_by_order[oid] = 0.0
            
            # Buscar costo del SKU
            costo_unitario = 0.0
            if sku:
                for fc_sku, cost in fc_costs_map.items():
                    if fc_sku in sku:
                        costo_unitario = cost
                        break
            
            costo_total_usd_by_order[oid] += costo_unitario * int(qty or 0)

        # 5. Combinar y calcular
        orders_data = []

        for row in orders_results:
            pedido_id, numero_pedido, fecha_pedido, estado, metodo_pago, total_venta_pen, costo_envio_pen, cliente_nombre, cliente_apellido = row
            
            total_venta_pen = float(total_venta_pen or 0)
            costo_envio_pen = float(costo_envio_pen or 0)
            tipo_cambio = get_tc_for_date(fecha_pedido)
            
            costo_total_usd = costo_total_usd_by_order.get(pedido_id, 0.0)
            costo_total_pen = costo_total_usd * tipo_cambio
            
            # Normalizar método de pago a nombre de plataforma
            plataforma = normalize_payment_method(metodo_pago, is_whatsapp=True)

            # Calcular comisión: 5% si Plataforma contiene "TARJETA"
            comision_pen = 0
            if 'TARJETA' in plataforma:
                comision_pen = round(total_venta_pen * 0.05, 2)

            # Recalcular ganancia
            ganancia_pen = total_venta_pen - costo_total_pen - costo_envio_pen - comision_pen
            margen_porcentaje = (ganancia_pen / total_venta_pen * 100) if total_venta_pen > 0 else 0

            cliente_completo = f"{cliente_nombre or ''} {cliente_apellido or ''}".strip() or 'Sin nombre'

            orders_data.append({
                'pedido_id': pedido_id,
                'numero_pedido': numero_pedido or str(pedido_id),
                'fecha_pedido': fecha_pedido,
                'estado': estado,
                'metodo_pago': plataforma,
                'total_venta_pen': total_venta_pen,
                'tipo_cambio': tipo_cambio,
                'costo_total_usd': costo_total_usd,
                'costo_total_pen': costo_total_pen,
                'comision_pen': comision_pen,
                'costo_envio_pen': costo_envio_pen,
                'ganancia_pen': ganancia_pen,
                'margen_porcentaje': margen_porcentaje,
                'cliente': cliente_completo or 'Sin nombre'
            })

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Pedidos Externos"

        # Estilos
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título del reporte
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = "Reporte de Ganancias - Pedidos Externos"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        # Período
        ws.merge_cells('A2:N2')
        period_cell = ws['A2']
        period_cell.value = f"Período: {start_date} a {end_date}"
        period_cell.alignment = Alignment(horizontal="center")

        # Headers (fila 4)
        headers = [
            'Pedido ID',
            'Número',
            'Fecha',
            'Estado',
            'Plataforma',
            'Venta (PEN)',
            'T.C.',
            'Costo (USD)',
            'Costo (PEN)',
            'Comisión (PEN)',
            'Envío (PEN)',
            'Ganancia (PEN)',
            'Margen %',
            'Cliente'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Datos
        row_num = 5
        for order in orders_data:
            ws.cell(row=row_num, column=1, value=order['pedido_id'])
            ws.cell(row=row_num, column=2, value=order['numero_pedido'])
            ws.cell(row=row_num, column=3, value=str(order['fecha_pedido']))
            ws.cell(row=row_num, column=4, value=order['estado'])
            ws.cell(row=row_num, column=5, value=order['metodo_pago'])
            ws.cell(row=row_num, column=6, value=round(order['total_venta_pen'], 2))
            ws.cell(row=row_num, column=7, value=round(order['tipo_cambio'], 2))
            ws.cell(row=row_num, column=8, value=round(order['costo_total_usd'], 2))
            ws.cell(row=row_num, column=9, value=round(order['costo_total_pen'], 2))
            ws.cell(row=row_num, column=10, value=round(order['comision_pen'], 2))
            ws.cell(row=row_num, column=11, value=round(order['costo_envio_pen'], 2))
            ws.cell(row=row_num, column=12, value=round(order['ganancia_pen'], 2))
            ws.cell(row=row_num, column=13, value=round(order['margen_porcentaje'], 2))
            ws.cell(row=row_num, column=14, value=order['cliente'])

            # Aplicar bordes
            for col in range(1, 15):
                ws.cell(row=row_num, column=col).border = border

            row_num += 1

        # Ajustar anchos de columna
        column_widths = [10, 15, 12, 12, 20, 12, 8, 12, 12, 12, 12, 14, 10, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Guardar en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Nombre del archivo
        filename = f"ganancias_externos_{start_date}_{end_date}.xlsx"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/api/profits/export')
@login_required
def export_profits_excel():
    """
    Exportar reporte de ganancias a Excel

    Query params:
    - start_date: fecha inicio (YYYY-MM-DD)
    - end_date: fecha fin (YYYY-MM-DD)
    - source: filtro de origen ('all', 'whatsapp', 'woocommerce') - default: 'all'
    - statuses: estados a incluir separados por coma (ej: 'wc-completed,wc-processing')
    """
    try:
        # Obtener parámetros (mismo código que api_profits)
        start_date = request.args.get('start_date', type=str)
        end_date = request.args.get('end_date', type=str)
        source = request.args.get('source', type=str, default='all')
        statuses_param = request.args.get('statuses', type=str, default='')

        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')

        # Procesar filtro de estados
        if statuses_param:
            # Convertir string de estados a lista y limpiar
            status_list = [s.strip() for s in statuses_param.split(',') if s.strip()]
            if status_list:
                # Crear filtro IN con los estados seleccionados
                status_placeholders = ', '.join([f"'{s}'" for s in status_list])
                status_filter = f"AND o.status IN ({status_placeholders})"
            else:
                # Por defecto excluir cancelados, reembolsados y fallidos
                status_filter = "AND o.status NOT IN ('wc-cancelled', 'wc-refunded', 'wc-failed')"
        else:
            # Por defecto excluir cancelados, reembolsados y fallidos
            status_filter = "AND o.status NOT IN ('wc-cancelled', 'wc-refunded', 'wc-failed')"

        # Determinar filtro de origen
        source_filter = ""
        source_name = "Todos"
        if source == 'whatsapp':
            source_filter = "AND (om_numero.meta_value IS NOT NULL OR oext.id IS NOT NULL)"
            source_name = "WhatsApp"
        elif source == 'woocommerce':
            source_filter = "AND om_numero.meta_value IS NULL AND oext.id IS NULL"
            source_name = "WooCommerce"

        # 1. Obtener todos los costos de FC en un diccionario para búsqueda rápida
        fc_costs_rows = db.session.execute(text("SELECT sku, FCLastCost FROM woo_products_fccost WHERE LENGTH(sku) = 7")).fetchall()
        fc_costs_map = {row[0]: float(row[1] or 0) for row in fc_costs_rows}

        # 2. Obtener tipos de cambio históricos
        tc_rows = db.session.execute(text("SELECT fecha, tasa_promedio FROM woo_tipo_cambio WHERE activo = TRUE ORDER BY fecha DESC")).fetchall()
        def get_tc_for_date(order_date):
            order_date_str = str(order_date)
            for tc_date, rate in tc_rows:
                if str(tc_date) <= order_date_str:
                    return float(rate)
            return 1.0

        # 3. Consulta de Órdenes (Solo datos básicos + plataforma)
        orders_sql = text(f"""
            SELECT
                o.id,
                COALESCE(om_numero.meta_value, CAST(o.id AS CHAR)) as numero_pedido,
                DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR)) as fecha,
                o.status,
                COALESCE(o.payment_method_title, o.payment_method, 'N/A') as metodo_pago,
                o.total_amount,
                CONCAT(ba.first_name, ' ', ba.last_name) as cliente_nombre,
                COALESCE((
                    SELECT SUM(CAST(oim_shipping.meta_value AS DECIMAL(10,2)))
                    FROM wpyz_woocommerce_order_items oi_shipping
                    INNER JOIN wpyz_woocommerce_order_itemmeta oim_shipping ON oi_shipping.order_item_id = oim_shipping.order_item_id
                    WHERE oi_shipping.order_id = o.id
                        AND oi_shipping.order_item_type = 'shipping'
                        AND oim_shipping.meta_key = 'cost'
                ), 0) as costo_envio
            FROM wpyz_wc_orders o
            LEFT JOIN wpyz_wc_orders_meta om_numero ON o.id = om_numero.order_id AND om_numero.meta_key = '_order_number'
            LEFT JOIN wpyz_wc_order_addresses ba ON o.id = ba.order_id AND ba.address_type = 'billing'
            LEFT JOIN woo_orders_ext oext ON om_numero.meta_value COLLATE utf8mb4_unicode_ci = oext.order_number
            WHERE DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR)) BETWEEN :start_date AND :end_date
                AND o.status != 'trash'
                {status_filter}
                {source_filter}
            ORDER BY fecha DESC, o.id DESC
        """)
        
        orders_results = db.session.execute(orders_sql, {'start_date': start_date, 'end_date': end_date}).fetchall()
        order_ids = [r[0] for r in orders_results]

        # 4. Consulta masiva de items para todas las órdenes encontradas
        items_by_order = {}
        if order_ids:
            items_sql = text("""
                SELECT
                    oi.order_id,
                    oim_qty.meta_value as qty,
                    pm_sku.meta_value as sku
                FROM wpyz_woocommerce_order_items oi
                INNER JOIN wpyz_woocommerce_order_itemmeta oim_pid ON oi.order_item_id = oim_pid.order_item_id AND oim_pid.meta_key = '_product_id'
                INNER JOIN wpyz_woocommerce_order_itemmeta oim_qty ON oi.order_item_id = oim_qty.order_item_id AND oim_qty.meta_key = '_qty'
                LEFT JOIN wpyz_woocommerce_order_itemmeta oim_vid ON oi.order_item_id = oim_vid.order_item_id AND oim_vid.meta_key = '_variation_id'
                LEFT JOIN wpyz_postmeta pm_sku ON CAST(COALESCE(NULLIF(oim_vid.meta_value, '0'), oim_pid.meta_value) AS UNSIGNED) = pm_sku.post_id AND pm_sku.meta_key = '_sku'
                WHERE oi.order_id IN :order_ids AND oi.order_item_type = 'line_item'
            """)
            items_results = db.session.execute(items_sql, {'order_ids': order_ids}).fetchall()
            
            for oid, qty, sku in items_results:
                if oid not in items_by_order: items_by_order[oid] = 0.0
                if sku:
                    for fc_sku, cost in fc_costs_map.items():
                        if fc_sku in sku:
                            items_by_order[oid] += float(cost) * int(qty or 0)
                            break

        # Crear workbook de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Ganancias"

        # Estilos etc (mismo código ...)
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = f"Reporte de Ganancias - {source_name}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        ws.merge_cells('A2:N2')
        period_cell = ws['A2']
        period_cell.value = f"Período: {start_date} a {end_date}"
        period_cell.alignment = Alignment(horizontal="center")

        headers = ['Pedido ID', 'Número', 'Fecha', 'Estado', 'Plataforma', 'Venta (PEN)', 'T.C.', 'Costo (USD)', 'Costo (PEN)', 'Comisión (PEN)', 'Envío (PEN)', 'Ganancia (PEN)', 'Margen %', 'Cliente']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, header_alignment, border

        # Datos
        row_num = 5
        for r in orders_results:
            oid, num, fecha, status, p_raw, total_venta_pen, cliente, c_envio = r
            tc = get_tc_for_date(fecha)
            costo_usd = items_by_order.get(oid, 0.0)
            costo_pen = costo_usd * tc
            envio_pen = float(c_envio or 0)

            # Normalizar plataforma
            is_whatsapp = num and str(num).startswith('W-')
            plataforma = normalize_payment_method(p_raw, is_whatsapp=is_whatsapp)

            comision_pen = round(float(total_venta_pen or 0) * 0.05, 2) if 'TARJETA' in plataforma else 0
            ganancia_pen = round(float(total_venta_pen or 0) - costo_pen - envio_pen - comision_pen, 2)
            margen_porcentaje = round((ganancia_pen / float(total_venta_pen or 1) * 100), 2) if total_venta_pen and float(total_venta_pen) > 0 else 0

            # Solo incluir pedidos con items con costo (HAVING anterior)
            if costo_usd == 0 and oid not in items_by_order: continue

            ws.cell(row=row_num, column=1, value=oid)
            ws.cell(row=row_num, column=2, value=num or oid)
            ws.cell(row=row_num, column=3, value=str(fecha))
            ws.cell(row=row_num, column=4, value=status)
            ws.cell(row=row_num, column=5, value=plataforma)
            ws.cell(row=row_num, column=6, value=float(total_venta_pen or 0))
            ws.cell(row=row_num, column=7, value=tc)
            ws.cell(row=row_num, column=8, value=costo_usd)
            ws.cell(row=row_num, column=9, value=costo_pen)
            ws.cell(row=row_num, column=10, value=comision_pen)
            ws.cell(row=row_num, column=11, value=envio_pen)
            ws.cell(row=row_num, column=12, value=ganancia_pen)
            ws.cell(row=row_num, column=13, value=margen_porcentaje)
            ws.cell(row=row_num, column=14, value=cliente or '')

            for col in range(1, 15):
                ws.cell(row=row_num, column=col).border = border
            row_num += 1

        # Ajustar anchos de columna
        column_widths = [10, 15, 12, 12, 20, 12, 8, 12, 12, 12, 12, 14, 10, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Guardar en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Nombre del archivo
        filename = f"ganancias_{source_name.lower()}_{start_date}_{end_date}.xlsx"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/exchange-rate')
@login_required
def exchange_rate():
    """
    Página de gestión de tipo de cambio
    """
    return render_template('exchange_rate.html', title='Tipo de Cambio USD/PEN')


@bp.route('/api/exchange-rate', methods=['GET'])
@login_required
def api_get_exchange_rate():
    """
    Obtener tipo de cambio actual o histórico

    Query params:
    - fecha: fecha específica (YYYY-MM-DD) - opcional
    - limit: cantidad de registros históricos (default: 30)
    """
    try:
        fecha_param = request.args.get('fecha', type=str)
        limit = request.args.get('limit', type=int, default=30)

        if fecha_param:
            # Obtener tipo de cambio para fecha específica
            fecha = datetime.strptime(fecha_param, '%Y-%m-%d').date()
            tipo_cambio = TipoCambio.get_tasa_por_fecha(fecha)

            if tipo_cambio:
                return jsonify({
                    'success': True,
                    'data': {
                        'id': tipo_cambio.id,
                        'fecha': tipo_cambio.fecha.isoformat(),
                        'tasa_compra': float(tipo_cambio.tasa_compra),
                        'tasa_venta': float(tipo_cambio.tasa_venta),
                        'tasa_promedio': float(tipo_cambio.tasa_promedio),
                        'actualizado_por': tipo_cambio.actualizado_por,
                        'notas': tipo_cambio.notas
                    }
                })
            else:
                return jsonify({
                    'success': False,
                    'error': 'No se encontró tipo de cambio para esa fecha'
                }), 404
        else:
            # Obtener histórico
            tipos_cambio = TipoCambio.query.filter_by(activo=True)\
                .order_by(TipoCambio.fecha.desc())\
                .limit(limit)\
                .all()

            data = [{
                'id': tc.id,
                'fecha': tc.fecha.isoformat(),
                'tasa_compra': float(tc.tasa_compra),
                'tasa_venta': float(tc.tasa_venta),
                'tasa_promedio': float(tc.tasa_promedio),
                'actualizado_por': tc.actualizado_por,
                'notas': tc.notas or ''
            } for tc in tipos_cambio]

            return jsonify({
                'success': True,
                'data': data
            })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/api/exchange-rate', methods=['POST'])
@login_required
def api_create_exchange_rate():
    """
    Crear o actualizar tipo de cambio

    Body JSON:
    {
        "fecha": "YYYY-MM-DD",
        "tasa_compra": 3.75,
        "tasa_venta": 3.78,
        "notas": "opcional"
    }
    """
    try:
        data = request.get_json()

        # Validar datos requeridos
        if not data.get('fecha') or not data.get('tasa_compra') or not data.get('tasa_venta'):
            return jsonify({
                'success': False,
                'error': 'Faltan campos requeridos: fecha, tasa_compra, tasa_venta'
            }), 400

        fecha = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
        tasa_compra = Decimal(str(data['tasa_compra']))
        tasa_venta = Decimal(str(data['tasa_venta']))
        tasa_promedio = (tasa_compra + tasa_venta) / 2
        notas = data.get('notas', '')

        # Verificar si ya existe un tipo de cambio para esa fecha
        existing = TipoCambio.query.filter_by(fecha=fecha, activo=True).first()

        if existing:
            # Actualizar existente
            existing.tasa_compra = tasa_compra
            existing.tasa_venta = tasa_venta
            existing.tasa_promedio = tasa_promedio
            existing.actualizado_por = current_user.username
            existing.fecha_actualizacion = datetime.now()
            existing.notas = notas
            tipo_cambio = existing
        else:
            # Crear nuevo
            tipo_cambio = TipoCambio(
                fecha=fecha,
                tasa_compra=tasa_compra,
                tasa_venta=tasa_venta,
                tasa_promedio=tasa_promedio,
                actualizado_por=current_user.username,
                activo=True,
                notas=notas
            )
            db.session.add(tipo_cambio)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Tipo de cambio guardado exitosamente',
            'data': {
                'id': tipo_cambio.id,
                'fecha': tipo_cambio.fecha.isoformat(),
                'tasa_compra': float(tipo_cambio.tasa_compra),
                'tasa_venta': float(tipo_cambio.tasa_venta),
                'tasa_promedio': float(tipo_cambio.tasa_promedio)
            }
        })

    except ValueError as e:
        return jsonify({
            'success': False,
            'error': 'Formato de datos inválido: ' + str(e)
        }), 400
    except Exception as e:
        import traceback
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


# ============================================================================
# DASHBOARD DE GRÁFICOS
# ============================================================================

@bp.route('/profits/dashboard')
@login_required
def profits_dashboard():
    """
    Dashboard de gráficos de ganancias
    """
    return render_template('reports_profits_dashboard.html', title='Dashboard de Ganancias')


@bp.route('/api/profits/charts/monthly', methods=['GET'])
@login_required
def api_profits_monthly():
    """
    Obtener datos de ganancias agrupados por mes

    Query params:
    - start_date: fecha inicio (YYYY-MM-DD)
    - end_date: fecha fin (YYYY-MM-DD)

    Retorna datos mensuales de: ventas, costos, ganancias, margen
    """
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        if not start_date or not end_date:
            return jsonify({
                'success': False,
                'error': 'Se requieren start_date y end_date'
            }), 400

        # Query para datos mensuales
        monthly_query = text("""
            SELECT
                DATE_FORMAT(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR), '%Y-%m') as mes,
                COUNT(DISTINCT o.id) as total_pedidos,
                ROUND(SUM(o.total_amount), 2) as ventas_totales_pen,
                ROUND(SUM(
                    (
                        SELECT SUM(
                            (
                                SELECT SUM(fc.FCLastCost)
                                FROM woo_products_fccost fc
                                WHERE pm_sku.meta_value COLLATE utf8mb4_unicode_520_ci LIKE CONCAT('%', fc.sku COLLATE utf8mb4_unicode_520_ci, '%')
                                    AND LENGTH(fc.sku) = 7
                            ) * oim_qty.meta_value
                        )
                        FROM wpyz_woocommerce_order_items oi
                        INNER JOIN wpyz_woocommerce_order_itemmeta oim_pid ON oi.order_item_id = oim_pid.order_item_id
                            AND oim_pid.meta_key = '_product_id'
                        INNER JOIN wpyz_woocommerce_order_itemmeta oim_qty ON oi.order_item_id = oim_qty.order_item_id
                            AND oim_qty.meta_key = '_qty'
                        LEFT JOIN wpyz_woocommerce_order_itemmeta oim_vid ON oi.order_item_id = oim_vid.order_item_id
                            AND oim_vid.meta_key = '_variation_id'
                        INNER JOIN wpyz_postmeta pm_sku ON CAST(COALESCE(NULLIF(oim_vid.meta_value, '0'), oim_pid.meta_value) AS UNSIGNED) = pm_sku.post_id
                            AND pm_sku.meta_key = '_sku'
                        WHERE oi.order_id = o.id
                            AND oi.order_item_type = 'line_item'
                    ) * (
                        SELECT tasa_promedio
                        FROM woo_tipo_cambio tc
                        WHERE tc.fecha <= DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR))
                            AND tc.activo = TRUE
                        ORDER BY tc.fecha DESC
                        LIMIT 1
                    )
                ), 2) as costos_totales_pen,
                ROUND(SUM(
                    COALESCE((
                        SELECT SUM(CAST(oim_shipping.meta_value AS DECIMAL(10,2)))
                        FROM wpyz_woocommerce_order_items oi_shipping
                        INNER JOIN wpyz_woocommerce_order_itemmeta oim_shipping ON oi_shipping.order_item_id = oim_shipping.order_item_id
                        WHERE oi_shipping.order_id = o.id
                            AND oi_shipping.order_item_type = 'shipping'
                            AND oim_shipping.meta_key = 'cost'
                    ), 0)
                ), 2) as costos_envio_totales_pen,
                ROUND(SUM(o.total_amount) - SUM(
                    (
                        SELECT SUM(
                            (
                                SELECT SUM(fc.FCLastCost)
                                FROM woo_products_fccost fc
                                WHERE pm_sku.meta_value COLLATE utf8mb4_unicode_520_ci LIKE CONCAT('%', fc.sku COLLATE utf8mb4_unicode_520_ci, '%')
                                    AND LENGTH(fc.sku) = 7
                            ) * oim_qty.meta_value
                        )
                        FROM wpyz_woocommerce_order_items oi
                        INNER JOIN wpyz_woocommerce_order_itemmeta oim_pid ON oi.order_item_id = oim_pid.order_item_id
                            AND oim_pid.meta_key = '_product_id'
                        INNER JOIN wpyz_woocommerce_order_itemmeta oim_qty ON oi.order_item_id = oim_qty.order_item_id
                            AND oim_qty.meta_key = '_qty'
                        LEFT JOIN wpyz_woocommerce_order_itemmeta oim_vid ON oi.order_item_id = oim_vid.order_item_id
                            AND oim_vid.meta_key = '_variation_id'
                        INNER JOIN wpyz_postmeta pm_sku ON CAST(COALESCE(NULLIF(oim_vid.meta_value, '0'), oim_pid.meta_value) AS UNSIGNED) = pm_sku.post_id
                            AND pm_sku.meta_key = '_sku'
                        WHERE oi.order_id = o.id
                            AND oi.order_item_type = 'line_item'
                    ) * (
                        SELECT tasa_promedio
                        FROM woo_tipo_cambio tc
                        WHERE tc.fecha <= DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR))
                            AND tc.activo = TRUE
                        ORDER BY tc.fecha DESC
                        LIMIT 1
                    )
                ) - SUM(
                    COALESCE((
                        SELECT SUM(CAST(oim_shipping.meta_value AS DECIMAL(10,2)))
                        FROM wpyz_woocommerce_order_items oi_shipping
                        INNER JOIN wpyz_woocommerce_order_itemmeta oim_shipping ON oi_shipping.order_item_id = oim_shipping.order_item_id
                        WHERE oi_shipping.order_id = o.id
                            AND oi_shipping.order_item_type = 'shipping'
                            AND oim_shipping.meta_key = 'cost'
                    ), 0)
                ), 2) as ganancias_totales_pen,
                ROUND(
                    (SUM(o.total_amount) - SUM(
                        (
                            SELECT SUM(
                                (
                                    SELECT SUM(fc.FCLastCost)
                                    FROM woo_products_fccost fc
                                    WHERE pm_sku.meta_value COLLATE utf8mb4_unicode_520_ci LIKE CONCAT('%', fc.sku COLLATE utf8mb4_unicode_520_ci, '%')
                                        AND LENGTH(fc.sku) = 7
                                ) * oim_qty.meta_value
                            )
                            FROM wpyz_woocommerce_order_items oi
                            INNER JOIN wpyz_woocommerce_order_itemmeta oim_pid ON oi.order_item_id = oim_pid.order_item_id
                                AND oim_pid.meta_key = '_product_id'
                            INNER JOIN wpyz_woocommerce_order_itemmeta oim_qty ON oi.order_item_id = oim_qty.order_item_id
                                AND oim_qty.meta_key = '_qty'
                            LEFT JOIN wpyz_woocommerce_order_itemmeta oim_vid ON oi.order_item_id = oim_vid.order_item_id
                                AND oim_vid.meta_key = '_variation_id'
                            INNER JOIN wpyz_postmeta pm_sku ON CAST(COALESCE(NULLIF(oim_vid.meta_value, '0'), oim_pid.meta_value) AS UNSIGNED) = pm_sku.post_id
                                AND pm_sku.meta_key = '_sku'
                            WHERE oi.order_id = o.id
                                AND oi.order_item_type = 'line_item'
                        ) * (
                            SELECT tasa_promedio
                            FROM woo_tipo_cambio tc
                            WHERE tc.fecha <= DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR))
                                AND tc.activo = TRUE
                            ORDER BY tc.fecha DESC
                            LIMIT 1
                        )
                    ) - SUM(
                        COALESCE((
                            SELECT SUM(CAST(oim_shipping.meta_value AS DECIMAL(10,2)))
                            FROM wpyz_woocommerce_order_items oi_shipping
                            INNER JOIN wpyz_woocommerce_order_itemmeta oim_shipping ON oi_shipping.order_item_id = oim_shipping.order_item_id
                            WHERE oi_shipping.order_id = o.id
                                AND oi_shipping.order_item_type = 'shipping'
                                AND oim_shipping.meta_key = 'cost'
                        ), 0)
                    )) / NULLIF(SUM(o.total_amount), 0) * 100
                , 2) as margen_promedio_porcentaje
            FROM wpyz_wc_orders o
            INNER JOIN wpyz_wc_orders_meta om_numero ON o.id = om_numero.order_id
                AND om_numero.meta_key = '_order_number'
            WHERE DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR)) BETWEEN :start_date AND :end_date
                AND o.status != 'trash'
                AND o.status NOT IN ('wc-cancelled', 'wc-refunded', 'wc-failed')
            GROUP BY mes
            ORDER BY mes
        """)

        results = db.session.execute(monthly_query, {
            'start_date': start_date,
            'end_date': end_date
        }).fetchall()

        monthly_data = []
        for row in results:
            monthly_data.append({
                'mes': row[0],
                'total_pedidos': row[1] or 0,
                'ventas_totales_pen': float(row[2] or 0),
                'costos_totales_pen': float(row[3] or 0),
                'costos_envio_totales_pen': float(row[4] or 0),
                'ganancias_totales_pen': float(row[5] or 0),
                'margen_promedio_porcentaje': float(row[6] or 0)
            })

        return jsonify({
            'success': True,
            'data': monthly_data
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/api/profits/charts/top-products', methods=['GET'])
@login_required
def api_profits_top_products():
    """
    Obtener top productos más rentables

    Query params:
    - start_date: fecha inicio (YYYY-MM-DD)
    - end_date: fecha fin (YYYY-MM-DD)
    - limit: cantidad de productos a retornar (default: 10)

    Retorna productos ordenados por ganancia total
    """
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        limit = int(request.args.get('limit', 10))

        if not start_date or not end_date:
            return jsonify({
                'success': False,
                'error': 'Se requieren start_date y end_date'
            }), 400

        # Query para top productos
        top_products_query = text("""
            SELECT
                oi.order_item_name as producto,
                SUM(oim_qty.meta_value) as cantidad_vendida,
                ROUND(SUM(oim_subtotal.meta_value), 2) as ventas_totales_pen,
                ROUND(SUM(
                    (
                        SELECT SUM(fc.FCLastCost)
                        FROM woo_products_fccost fc
                        WHERE pm_sku.meta_value COLLATE utf8mb4_unicode_520_ci LIKE CONCAT('%', fc.sku COLLATE utf8mb4_unicode_520_ci, '%')
                            AND LENGTH(fc.sku) = 7
                    ) * oim_qty.meta_value * (
                        SELECT tasa_promedio
                        FROM woo_tipo_cambio tc
                        WHERE tc.fecha <= DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR))
                            AND tc.activo = TRUE
                        ORDER BY tc.fecha DESC
                        LIMIT 1
                    )
                ), 2) as costos_totales_pen,
                ROUND(SUM(oim_subtotal.meta_value) - SUM(
                    (
                        SELECT SUM(fc.FCLastCost)
                        FROM woo_products_fccost fc
                        WHERE pm_sku.meta_value COLLATE utf8mb4_unicode_520_ci LIKE CONCAT('%', fc.sku COLLATE utf8mb4_unicode_520_ci, '%')
                            AND LENGTH(fc.sku) = 7
                    ) * oim_qty.meta_value * (
                        SELECT tasa_promedio
                        FROM woo_tipo_cambio tc
                        WHERE tc.fecha <= DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR))
                            AND tc.activo = TRUE
                        ORDER BY tc.fecha DESC
                        LIMIT 1
                    )
                ), 2) as ganancia_total_pen,
                ROUND(
                    (SUM(oim_subtotal.meta_value) - SUM(
                        (
                            SELECT SUM(fc.FCLastCost)
                            FROM woo_products_fccost fc
                            WHERE pm_sku.meta_value COLLATE utf8mb4_unicode_520_ci LIKE CONCAT('%', fc.sku COLLATE utf8mb4_unicode_520_ci, '%')
                                AND LENGTH(fc.sku) = 7
                        ) * oim_qty.meta_value * (
                            SELECT tasa_promedio
                            FROM woo_tipo_cambio tc
                            WHERE tc.fecha <= DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR))
                                AND tc.activo = TRUE
                            ORDER BY tc.fecha DESC
                            LIMIT 1
                        )
                    )) / NULLIF(SUM(oim_subtotal.meta_value), 0) * 100
                , 2) as margen_porcentaje
            FROM wpyz_woocommerce_order_items oi
            INNER JOIN wpyz_wc_orders o ON oi.order_id = o.id
            INNER JOIN wpyz_woocommerce_order_itemmeta oim_pid ON oi.order_item_id = oim_pid.order_item_id
                AND oim_pid.meta_key = '_product_id'
            INNER JOIN wpyz_woocommerce_order_itemmeta oim_qty ON oi.order_item_id = oim_qty.order_item_id
                AND oim_qty.meta_key = '_qty'
            INNER JOIN wpyz_woocommerce_order_itemmeta oim_subtotal ON oi.order_item_id = oim_subtotal.order_item_id
                AND oim_subtotal.meta_key = '_line_subtotal'
            LEFT JOIN wpyz_woocommerce_order_itemmeta oim_vid ON oi.order_item_id = oim_vid.order_item_id
                AND oim_vid.meta_key = '_variation_id'
            LEFT JOIN wpyz_postmeta pm_sku ON CAST(COALESCE(NULLIF(oim_vid.meta_value, '0'), oim_pid.meta_value) AS UNSIGNED) = pm_sku.post_id
                AND pm_sku.meta_key = '_sku'
            WHERE DATE(DATE_SUB(o.date_created_gmt, INTERVAL 5 HOUR)) BETWEEN :start_date AND :end_date
                AND oi.order_item_type = 'line_item'
                AND o.status != 'trash'
                AND o.status NOT IN ('wc-cancelled', 'wc-refunded', 'wc-failed')
            GROUP BY oi.order_item_name
            HAVING ganancia_total_pen IS NOT NULL
            ORDER BY ganancia_total_pen DESC
            LIMIT :limit
        """)

        results = db.session.execute(top_products_query, {
            'start_date': start_date,
            'end_date': end_date,
            'limit': limit
        }).fetchall()

        products_data = []
        for row in results:
            products_data.append({
                'producto': row[0],
                'cantidad_vendida': int(row[1] or 0),
                'ventas_totales_pen': float(row[2] or 0),
                'costos_totales_pen': float(row[3] or 0),
                'ganancia_total_pen': float(row[4] or 0),
                'margen_porcentaje': float(row[5] or 0)
            })

        return jsonify({
            'success': True,
            'data': products_data
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@bp.route('/api/profits/charts/by-advisor', methods=['GET'])
@login_required
def api_profits_by_advisor():
    """
    API para obtener ganancias por asesor
    """
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        if not start_date or not end_date:
            return jsonify({
                'success': False,
                'error': 'Se requieren start_date y end_date'
            }), 400

        # Query para obtener datos por asesor (usuario que creó el pedido)
        query = text("""
            SELECT
                COALESCE(pm_created.meta_value, 'WooCommerce') as asesor_nombre,
                COUNT(DISTINCT o.id) as total_pedidos,
                COALESCE(SUM(o.total_amount), 0) as ventas_totales_pen,
                COALESCE(SUM(
                    (SELECT SUM(fc.FCLastCost * CAST(oim_qty.meta_value AS DECIMAL(10,2)))
                    FROM wpyz_woocommerce_order_items oi
                    INNER JOIN wpyz_woocommerce_order_itemmeta oim_pid ON oi.order_item_id = oim_pid.order_item_id
                        AND oim_pid.meta_key = '_product_id'
                    LEFT JOIN wpyz_woocommerce_order_itemmeta oim_vid ON oi.order_item_id = oim_vid.order_item_id
                        AND oim_vid.meta_key = '_variation_id'
                    INNER JOIN wpyz_postmeta pm_sku ON CAST(COALESCE(NULLIF(oim_vid.meta_value, '0'), oim_pid.meta_value) AS UNSIGNED) = pm_sku.post_id
                        AND pm_sku.meta_key = '_sku'
                    INNER JOIN woo_products_fccost fc ON pm_sku.meta_value COLLATE utf8mb4_unicode_520_ci LIKE CONCAT('%', fc.sku COLLATE utf8mb4_unicode_520_ci, '%')
                        AND LENGTH(fc.sku) = 7
                    INNER JOIN wpyz_woocommerce_order_itemmeta oim_qty ON oi.order_item_id = oim_qty.order_item_id
                        AND oim_qty.meta_key = '_qty'
                    WHERE oi.order_id = o.id
                        AND oi.order_item_type = 'line_item')
                ), 0) as costos_totales_usd,
                COALESCE(tc.tasa_promedio, 3.75) as tipo_cambio_promedio,
                COALESCE(SUM(
                    (SELECT SUM(CAST(oim_shipping.meta_value AS DECIMAL(10,2)))
                    FROM wpyz_woocommerce_order_items oi_shipping
                    INNER JOIN wpyz_woocommerce_order_itemmeta oim_shipping ON oi_shipping.order_item_id = oim_shipping.order_item_id
                    WHERE oi_shipping.order_id = o.id
                        AND oi_shipping.order_item_type = 'shipping'
                        AND oim_shipping.meta_key = 'cost')
                ), 0) as costos_envio_totales_pen
            FROM wpyz_wc_orders o
            LEFT JOIN wpyz_wc_orders_meta pm_created ON o.id = pm_created.order_id
                AND pm_created.meta_key = '_created_by'
            LEFT JOIN woo_tipo_cambio tc ON DATE(o.date_created_gmt) = tc.fecha
                AND tc.activo = 1
            WHERE DATE(o.date_created_gmt) BETWEEN :start_date AND :end_date
                AND o.status IN ('wc-completed', 'wc-processing')
            GROUP BY asesor_nombre
            HAVING ventas_totales_pen > 0
            ORDER BY ventas_totales_pen DESC
        """)

        result = db.session.execute(query, {
            'start_date': start_date,
            'end_date': end_date
        }).fetchall()

        advisors_data = []
        for row in result:
            costos_totales_pen = float(row[3] or 0) * float(row[4] or 3.75)
            costos_envio_totales_pen = float(row[5] or 0)
            ventas_totales_pen = float(row[2] or 0)
            ganancia_total_pen = ventas_totales_pen - costos_totales_pen - costos_envio_totales_pen
            margen_porcentaje = (ganancia_total_pen / ventas_totales_pen * 100) if ventas_totales_pen > 0 else 0

            advisors_data.append({
                'asesor_nombre': row[0],
                'total_pedidos': int(row[1] or 0),
                'ventas_totales_pen': ventas_totales_pen,
                'costos_totales_pen': costos_totales_pen,
                'costos_envio_totales_pen': costos_envio_totales_pen,
                'ganancia_total_pen': ganancia_total_pen,
                'margen_porcentaje': margen_porcentaje
            })

        return jsonify({
            'success': True,
            'data': advisors_data
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/api/profits/charts/by-status', methods=['GET'])
@login_required
def api_profits_by_status():
    """
    API para obtener distribución de pedidos por estado
    """
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        if not start_date or not end_date:
            return jsonify({
                'success': False,
                'error': 'Se requieren start_date y end_date'
            }), 400

        # Query para obtener datos por estado
        query = text("""
            SELECT
                o.status,
                COUNT(DISTINCT o.id) as total_pedidos,
                COALESCE(SUM(o.total_amount), 0) as ventas_totales_pen
            FROM wpyz_wc_orders o
            WHERE DATE(o.date_created_gmt) BETWEEN :start_date AND :end_date
            GROUP BY o.status
            ORDER BY total_pedidos DESC
        """)

        result = db.session.execute(query, {
            'start_date': start_date,
            'end_date': end_date
        }).fetchall()

        status_data = []
        for row in result:
            status_data.append({
                'status': row[0],
                'total_pedidos': int(row[1] or 0),
                'ventas_totales_pen': float(row[2] or 0)
            })

        return jsonify({
            'success': True,
            'data': status_data
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/api/profits/charts/low-margin-products', methods=['GET'])
@login_required
def api_profits_low_margin_products():
    """
    API para obtener productos con márgenes bajos o pérdidas
    """
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        threshold = float(request.args.get('threshold', 15))  # Margen mínimo aceptable (%)
        limit = int(request.args.get('limit', 10))

        if not start_date or not end_date:
            return jsonify({
                'success': False,
                'error': 'Se requieren start_date y end_date'
            }), 400

        # Query para productos con margen bajo (usando subconsulta para evitar error de MySQL)
        query = text("""
            SELECT
                producto,
                total_pedidos,
                cantidad_total,
                ventas_totales_pen,
                costos_totales_usd,
                tipo_cambio_promedio,
                (ventas_totales_pen - (costos_totales_usd * tipo_cambio_promedio)) as ganancia_pen,
                ((ventas_totales_pen - (costos_totales_usd * tipo_cambio_promedio)) / ventas_totales_pen * 100) as margen_porcentaje
            FROM (
                SELECT
                    oi.order_item_name as producto,
                    COUNT(DISTINCT o.id) as total_pedidos,
                    SUM(CAST(oim_qty.meta_value AS DECIMAL(10,2))) as cantidad_total,
                    COALESCE(SUM(CAST(oim_total.meta_value AS DECIMAL(10,2))), 0) as ventas_totales_pen,
                    COALESCE(SUM(
                        (SELECT SUM(fc.FCLastCost * CAST(oim_qty.meta_value AS DECIMAL(10,2)))
                        FROM woo_products_fccost fc
                        INNER JOIN wpyz_postmeta pm_sku ON pm_sku.meta_value COLLATE utf8mb4_unicode_520_ci LIKE CONCAT('%', fc.sku COLLATE utf8mb4_unicode_520_ci, '%')
                            AND LENGTH(fc.sku) = 7
                        WHERE pm_sku.post_id = CAST(COALESCE(NULLIF(oim_vid.meta_value, '0'), oim_pid.meta_value) AS UNSIGNED)
                            AND pm_sku.meta_key = '_sku')
                    ), 0) as costos_totales_usd,
                    AVG(COALESCE(tc.tasa_promedio, 3.75)) as tipo_cambio_promedio
                FROM wpyz_wc_orders o
                INNER JOIN wpyz_woocommerce_order_items oi ON o.id = oi.order_id
                    AND oi.order_item_type = 'line_item'
                INNER JOIN wpyz_woocommerce_order_itemmeta oim_pid ON oi.order_item_id = oim_pid.order_item_id
                    AND oim_pid.meta_key = '_product_id'
                LEFT JOIN wpyz_woocommerce_order_itemmeta oim_vid ON oi.order_item_id = oim_vid.order_item_id
                    AND oim_vid.meta_key = '_variation_id'
                INNER JOIN wpyz_woocommerce_order_itemmeta oim_qty ON oi.order_item_id = oim_qty.order_item_id
                    AND oim_qty.meta_key = '_qty'
                INNER JOIN wpyz_woocommerce_order_itemmeta oim_total ON oi.order_item_id = oim_total.order_item_id
                    AND oim_total.meta_key = '_line_total'
                LEFT JOIN woo_tipo_cambio tc ON DATE(o.date_created_gmt) = tc.fecha
                    AND tc.activo = 1
                WHERE DATE(o.date_created_gmt) BETWEEN :start_date AND :end_date
                    AND o.status IN ('wc-completed', 'wc-processing')
                GROUP BY producto
            ) AS productos_agregados
            WHERE ventas_totales_pen > 0
            ORDER BY margen_porcentaje ASC
            LIMIT :limit
        """)

        result = db.session.execute(query, {
            'start_date': start_date,
            'end_date': end_date,
            'limit': limit
        }).fetchall()

        products_data = []
        for row in result:
            # La subconsulta ya calcula ganancia_pen y margen_porcentaje
            costos_totales_pen = float(row[4] or 0) * float(row[5] or 3.75)
            ventas_totales_pen = float(row[3] or 0)
            ganancia_total_pen = float(row[6] or 0)  # Ya calculado en la query
            margen_porcentaje = float(row[7] or 0)   # Ya calculado en la query

            # Solo incluir si el margen es menor al umbral
            if margen_porcentaje < threshold:
                products_data.append({
                    'producto': row[0],
                    'total_pedidos': int(row[1] or 0),
                    'cantidad_total': float(row[2] or 0),
                    'ventas_totales_pen': ventas_totales_pen,
                    'costos_totales_pen': costos_totales_pen,
                    'ganancia_total_pen': ganancia_total_pen,
                    'margen_porcentaje': margen_porcentaje
                })

        return jsonify({
            'success': True,
            'data': products_data
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500
