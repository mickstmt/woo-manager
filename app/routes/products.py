# app/routes/products.py
from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash
from flask_login import login_required
from app.models import Product, ProductMeta, Term
from app import db, cache
from sqlalchemy import or_

# Crear el blueprint
bp = Blueprint('products', __name__, url_prefix='/products')


@bp.route('/')
@login_required
def index():
    """
    Ruta principal de productos
    Muestra la lista de todos los productos con paginación
    
    URL: http://localhost:5000/products/
    """
    return render_template('products.html', title='Gestión de Productos')


@bp.route('/list')
@login_required
@cache.cached(timeout=180, query_string=True)  # Cache 3 min basado en parámetros URL
def list_products():
    """
    Ruta para obtener lista de productos en formato JSON con paginación
    
    Comportamiento simple:
    - Sin búsqueda: Muestra últimos productos padre creados
    - Con búsqueda: Busca en productos padre Y variaciones (muestra padre si encuentra variación)
    """
    try:
        from sqlalchemy import text
        
        # Obtener parámetros de la URL
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        search = request.args.get('search', '', type=str)
        status = request.args.get('status', '', type=str)
        
        # Limitar per_page
        per_page = min(per_page, 500)
        
        # ========================================
        # CASO 1: Con búsqueda
        # ========================================
        if search and search.strip() != '':
            status_filter = f"AND p.post_status = '{status}'" if status else "AND p.post_status = 'publish'"
            
            # Separar términos de búsqueda (Keyword-based search)
            search_terms = search.strip().split()
            
            # Construir condiciones para búsqueda de productos padre (AND logic)
            parent_conditions = []
            params = {}
            for i, term in enumerate(search_terms):
                param_name = f'term_{i}'

                # Reglas de búsqueda inteligente:
                # - Si contiene guión y empieza con 7 dígitos: Buscar por SKU completo (ej: 1003228-CSHWGT4BR22M)
                # - Si son 5 dígitos exactos: Buscar por ID
                # - Si son 7 dígitos exactos: Buscar por SKU base (ej: 1003226 → 1003226-VARIANTE)
                # - Si no: Buscar solo por Título

                is_numeric = term.isdigit()
                term_len = len(term)
                has_dash = '-' in term

                # Detectar SKU completo con guión (ej: 1003228-CSHWGT4BR22M)
                if has_dash:
                    parts = term.split('-')
                    if len(parts) >= 2 and parts[0].isdigit() and len(parts[0]) == 7:
                        # Búsqueda por SKU completo
                        parent_conditions.append(f"(pm.meta_value LIKE :{param_name})")
                        params[param_name] = f'%{term}%'
                    else:
                        # Si tiene guión pero no coincide con formato SKU, buscar por título
                        parent_conditions.append(f"(p.post_title LIKE :{param_name})")
                        params[param_name] = f'%{term}%'
                elif is_numeric and term_len == 5:
                    # Búsqueda por ID exacto
                    parent_conditions.append(f"(p.ID = :{param_name})")
                    params[param_name] = term
                elif is_numeric and term_len == 7:
                    # Búsqueda por SKU base (encuentra 1003226 y 1003226-VARIANTE)
                    parent_conditions.append(f"(pm.meta_value LIKE :{param_name})")
                    params[param_name] = f'{term}%'
                else:
                    # Búsqueda por Título (parcial)
                    parent_conditions.append(f"(p.post_title LIKE :{param_name})")
                    params[param_name] = f'%{term}%'
            
            parent_where_clause = " AND ".join(parent_conditions)
            
            # Buscar en productos padre
            parent_query = text(f"""
                SELECT DISTINCT p.ID
                FROM wpyz_posts p
                LEFT JOIN wpyz_postmeta pm ON p.ID = pm.post_id AND pm.meta_key = '_sku'
                WHERE p.post_type = 'product'
                {status_filter}
                AND {parent_where_clause}
            """)
            
            parent_result = db.session.execute(parent_query, params)
            parent_ids = [row[0] for row in parent_result]
            
            # Construir condiciones para búsqueda de variaciones (AND logic)
            variation_conditions = []
            for i, term in enumerate(search_terms):
                param_name = f'term_{i}' # Los términos ya están en params

                is_numeric = term.isdigit()
                term_len = len(term)
                has_dash = '-' in term

                # Detectar SKU completo con guión (ej: 1003228-CSHWGT4BR22M)
                if has_dash:
                    parts = term.split('-')
                    if len(parts) >= 2 and parts[0].isdigit() and len(parts[0]) == 7:
                        # Búsqueda por SKU completo de variación
                        variation_conditions.append(f"(vm.meta_value LIKE :{param_name})")
                    else:
                        # Si tiene guión pero no coincide con formato SKU, buscar por título
                        variation_conditions.append(f"(v.post_title LIKE :{param_name})")
                elif is_numeric and term_len == 5:
                    # Si es 5 dígitos, buscar por ID de variación O ID padre
                    variation_conditions.append(f"(v.ID = :{param_name} OR v.post_parent = :{param_name})")
                elif is_numeric and term_len == 7:
                    # Si es 7 dígitos, buscar por SKU base de variación (encuentra 1003226 y 1003226-VARIANTE)
                    variation_conditions.append(f"(vm.meta_value LIKE :{param_name})")
                else:
                    # Búsqueda por título de variación
                    variation_conditions.append(f"(v.post_title LIKE :{param_name})")
            
            variation_where_clause = " AND ".join(variation_conditions)
            
            # Buscar en variaciones
            variation_query = text(f"""
                SELECT DISTINCT v.post_parent
                FROM wpyz_posts v
                LEFT JOIN wpyz_postmeta vm ON v.ID = vm.post_id AND vm.meta_key = '_sku'
                WHERE v.post_type = 'product_variation'
                AND v.post_parent != 0
                AND {variation_where_clause}
            """)
            
            variation_result = db.session.execute(variation_query, params)
            parent_ids_from_variations = [row[0] for row in variation_result if row[0]]
            
            # Combinar todos los IDs de productos padre (sin duplicados)
            all_parent_ids = list(set(parent_ids + parent_ids_from_variations))
            
            if not all_parent_ids:
                # No se encontró nada
                return jsonify({
                    'success': True,
                    'products': [],
                    'search_term': search,
                    'pagination': {
                        'page': page,
                        'per_page': per_page,
                        'total': 0,
                        'pages': 1,
                        'has_prev': False,
                        'has_next': False,
                        'prev_num': None,
                        'next_num': None
                    }
                })
            
            # Obtener productos padre
            total_products = len(all_parent_ids)
            
            # Aplicar paginación manualmente
            start_idx = (page - 1) * per_page
            end_idx = start_idx + per_page
            paginated_ids = all_parent_ids[start_idx:end_idx]
            
            products_items = Product.query.filter(Product.ID.in_(paginated_ids)).order_by(Product.ID.desc()).all()
            
            # Calcular páginas
            import math
            total_pages = math.ceil(total_products / per_page) if total_products > 0 else 1
            has_prev = page > 1
            has_next = page < total_pages
            prev_num = page - 1 if has_prev else None
            next_num = page + 1 if has_next else None
            
        # ========================================
        # CASO 2: Sin búsqueda (mostrar últimos productos padre)
        # ========================================
        else:
            query = Product.query.filter_by(post_type='product')
            
            if status:
                query = query.filter_by(post_status=status)
            else:
                query = query.filter_by(post_status='publish')
            
            total_products = query.count()
            
            # Ordenar por fecha de creación (más recientes primero)
            products = query.order_by(Product.post_date.desc()).paginate(
                page=page,
                per_page=per_page,
                error_out=False
            )
            
            products_items = products.items
            total_pages = products.pages
            has_prev = products.has_prev
            has_next = products.has_next
            prev_num = products.prev_num
            next_num = products.next_num
        
        # ========================================
        # Procesar productos padre - OPTIMIZADO
        # ========================================
        # Pre-cargar metadatos e imágenes para todos los productos en bulk
        Product.preload_metadata_for_products(products_items)
        Product.preload_images_for_products(products_items)

        # Contar variaciones de todos los productos en UNA SOLA consulta
        from sqlalchemy import func
        product_ids = [p.ID for p in products_items]
        variations_query = db.session.query(
            Product.post_parent,
            func.count(Product.ID).label('count')
        ).filter(
            Product.post_type == 'product_variation',
            Product.post_parent.in_(product_ids)
        ).group_by(Product.post_parent).all()

        # Crear diccionario para búsqueda rápida
        variations_dict = {row[0]: row[1] for row in variations_query}

        # Convertir a formato JSON
        products_list = []
        for product in products_items:
            # Obtener metadatos de la caché interna del objeto
            sku = product.get_meta('_sku') or 'N/A'
            price = product.get_meta('_price') or '0'
            stock = product.get_meta('_stock') or 'N/A'
            stock_status = product.get_meta('_stock_status') or 'instock'
            
            # Obtener cantidad de variaciones del diccionario
            variations_count = variations_dict.get(product.ID, 0)

            # Determinar tipo de producto
            is_variable = variations_count > 0
            product_type = 'variable' if is_variable else 'simple'

            # Obtener URL de imagen (desde la caché del objeto)
            image_url = product.get_image_url()

            products_list.append({
                'id': product.ID,
                'title': product.post_title,
                'status': product.post_status,
                'type': product_type,
                'sku': sku,
                'price': price,
                'stock': stock,
                'stock_status': stock_status,
                'is_variable': is_variable,
                'variations_count': variations_count,
                'image_url': image_url,
                'date': product.post_date.strftime('%Y-%m-%d %H:%M:%S') if product.post_date else None
            })
        
        return jsonify({
            'success': True,
            'products': products_list,
            'search_term': search if search else '',
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total_products,
                'pages': total_pages,
                'has_prev': has_prev,
                'has_next': has_next,
                'prev_num': prev_num,
                'next_num': next_num
            }
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/stats')
@login_required
def stats():
    """
    Obtener estadísticas generales de productos
    
    URL: http://localhost:5000/products/stats
    """
    try:
        # Total de productos
        total = Product.query.filter_by(post_type='product').count()
        
        # Productos publicados
        published = Product.query.filter_by(
            post_type='product',
            post_status='publish'
        ).count()
        
        # Productos en borrador
        draft = Product.query.filter_by(
            post_type='product',
            post_status='draft'
        ).count()
        
        # Productos pendientes
        pending = Product.query.filter_by(
            post_type='product',
            post_status='pending'
        ).count()
        
        return jsonify({
            'success': True,
            'stats': {
                'total': total,
                'published': published,
                'draft': draft,
                'pending': pending
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@bp.route('/<int:product_id>/variations')
@login_required
def get_variations(product_id):
    """
    Obtener todas las variaciones de un producto variable
    
    URL: http://localhost:5000/products/123/variations
    """
    try:
        # Buscar variaciones (productos hijos)
        variations = Product.query.filter_by(
            post_type='product_variation',
            post_parent=product_id
        ).order_by(Product.ID.asc()).all()
        
        # OPTIMIZACIÓN: Pre-cargar metadatos e imágenes para todas las variaciones de una vez
        Product.preload_metadata_for_products(variations)
        Product.preload_images_for_products(variations)
        
        variations_list = []
        for variation in variations:
            # Obtener atributos de la caché de metadatos (ya cargados)
            attributes = {}
            if hasattr(variation, '_meta_cache'):
                for key, value in variation._meta_cache.items():
                    if key.startswith('attribute_'):
                        if key.startswith('attribute_pa_'):
                            attr_name = key.replace('attribute_pa_', '')
                        else:
                            attr_name = key.replace('attribute_', '')
                        
                        if value:
                            attributes[attr_name] = value
            
            # Obtener datos importantes de la caché
            sku = variation.get_meta('_sku') or 'N/A'
            price = variation.get_meta('_price') or '0'
            regular_price = variation.get_meta('_regular_price') or '0'
            sale_price = variation.get_meta('_sale_price') or ''
            stock = variation.get_meta('_stock') or 'N/A'
            stock_status = variation.get_meta('_stock_status') or 'instock'

            # Obtener imagen de la variación (usa caché para thumbnail_id y get_image_url optimizado)
            image_url = variation.get_image_url()

            variations_list.append({
                'id': variation.ID,
                'title': variation.post_title,
                'attributes': attributes,
                'sku': sku,
                'price': price,
                'regular_price': regular_price,
                'sale_price': sale_price,
                'stock': stock,
                'stock_status': stock_status,
                'image_url': image_url
            })
        
        return jsonify({
            'success': True,
            'variations': variations_list,
            'total': len(variations_list)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/<int:product_id>')
@login_required
def view_product(product_id):
    """
    Ver detalles completos de un producto
    
    URL: /products/123
    """
    try:
        # Obtener el producto
        product = Product.query.get(product_id)
        
        if not product:
            flash('Producto no encontrado.', 'danger')
            return redirect(url_for('products.index'))
        
        # Obtener variaciones si es producto variable
        variations_query = []
        if product.post_type == 'product':
            variations_query = Product.query.filter_by(
                post_type='product_variation',
                post_parent=product.ID
            ).all()
        
        # OPTIMIZACIÓN: Pre-cargar metadatos e imágenes para el producto y todas sus variaciones en bulk
        all_products_to_load = [product] + variations_query
        Product.preload_metadata_for_products(all_products_to_load)
        Product.preload_images_for_products(all_products_to_load)
        
        # Obtener metadatos importantes del objeto ya cargado (usa caché interna)
        product_data = {
            'id': product.ID,
            'title': product.post_title,
            'status': product.post_status,
            'type': product.post_type,
            'description': product.post_content,
            'short_description': product.post_excerpt,
            'created_at': product.post_date,
            'modified_at': product.post_modified,
            
            # Metadatos (ahora vienen de caché)
            'sku': product.get_meta('_sku') or 'N/A',
            'regular_price': product.get_meta('_regular_price') or '0',
            'sale_price': product.get_meta('_sale_price') or '',
            'price': product.get_meta('_price') or '0',
            'stock': product.get_meta('_stock') or '0',
            'stock_status': product.get_meta('_stock_status') or 'outofstock',
            'manage_stock': product.get_meta('_manage_stock') or 'no',
            'backorders': product.get_meta('_backorders') or 'no',
            'sold_individually': product.get_meta('_sold_individually') or 'no',
            
            # Dimensiones
            'weight': product.get_meta('_weight') or '',
            'length': product.get_meta('_length') or '',
            'width': product.get_meta('_width') or '',
            'height': product.get_meta('_height') or '',
            
            # Imagen
            'image_url': product.get_image_url(),
            
            # Producto padre (si es variación)
            'parent_id': product.post_parent
        }
        
        # Si es variación, obtener el producto padre
        parent_product = None
        if product.post_type == 'product_variation' and product.post_parent:
            parent_product = Product.query.get(product.post_parent)
            if parent_product:
                parent_product.preload_meta() # Cargar sus metas también
        
        # --- PROCESAMIENTO DE VARIACIONES (AGRUPACIÓN MULTINIVEL) ---
        # 1. Recolectar todos los slugs de atributos globales para buscar sus nombres reales
        term_slugs = set()
        for var in variations_query:
            # Obtener metas directamente de la caché pre-cargada
            if hasattr(var, '_meta_cache'):
                for key, value in var._meta_cache.items():
                    if key.startswith('attribute_pa_') and value:
                        term_slugs.add(value)
        
        term_map = {}
        if term_slugs:
            terms = Term.query.filter(Term.slug.in_(term_slugs)).all()
            term_map = {t.slug: t.name for t in terms}

        # 2. Procesar y agrupar variaciones
        # Estructura: { "Medida": { "Conector": [variaciones] } }
        variations_grouped = {}
        processed_variations = []

        for var in variations_query:
            # Extraer todos los atributos de forma estructurada
            attr_list = []
            medida_val = 'General'
            conector_val = 'Otros'

            if hasattr(var, '_meta_cache'):
                for key, value in var._meta_cache.items():
                    if key.startswith('attribute_'):
                        attr_slug = value
                        is_global = key.startswith('attribute_pa_')
                        
                        # Valor por defecto
                        display_val = attr_slug if attr_slug else "Cualquiera"
                        if is_global and attr_slug in term_map:
                            display_val = term_map[attr_slug]
                        
                        # Limpiar nombre del atributo
                        attr_name = key.replace('attribute_pa_', '').replace('attribute_', '').replace('_', ' ').replace('-', ' ').title()
                        
                        # Normalizar capitalización (especialmente para mm)
                        if display_val and not any(char.isdigit() for char in display_val):
                            display_val = display_val.title()
                        
                        attr_list.append({'name': attr_name, 'value': display_val})

                        # Identificar claves de agrupación
                        if 'medida' in attr_name.lower():
                            medida_val = display_val
                        elif 'conector' in attr_name.lower():
                            conector_val = display_val

            var_data = {
                'id': var.ID,
                'title': var.post_title,
                'sku': var.get_meta('_sku') or 'N/A',
                'price': var.get_meta('_price') or '0',
                'stock': var.get_meta('_stock') or '0',
                'stock_status': var.get_meta('_stock_status') or 'outofstock',
                'image_url': var.get_image_url(),
                'attributes': attr_list,
                'medida': medida_val,
                'conector': conector_val
            }

            processed_variations.append(var_data)

            # Agrupar
            if medida_val not in variations_grouped:
                variations_grouped[medida_val] = {}
            if conector_val not in variations_grouped[medida_val]:
                variations_grouped[medida_val][conector_val] = []
            variations_grouped[medida_val][conector_val].append(var_data)
        
        # Ordenar los grupos para consistencia en la UI
        # (Convertir a lista de tuplas ordenadas si es necesario, o manejar en template)
        
        # Obtener categorías
        from sqlalchemy import text
        categories_query = text("""
            SELECT t.name, t.slug
            FROM wpyz_term_relationships tr
            JOIN wpyz_term_taxonomy tt ON tr.term_taxonomy_id = tt.term_taxonomy_id
            JOIN wpyz_terms t ON tt.term_id = t.term_id
            WHERE tr.object_id = :product_id
            AND tt.taxonomy = 'product_cat'
        """)
        
        categories_result = db.session.execute(categories_query, {'product_id': product_id})
        categories = [{'name': row[0], 'slug': row[1]} for row in categories_result]
        
        # Obtener historial de stock
        from app.models import StockHistory
        stock_history = StockHistory.query.filter_by(
            product_id=product_id
        ).order_by(StockHistory.created_at.desc()).limit(10).all()
        
        return render_template(
            'products/view.html',
            product=product_data,
            parent_product=parent_product,
            variations=processed_variations,
            variations_grouped=variations_grouped,
            categories=categories,
            stock_history=stock_history
        )
        
    except Exception as e:
        import traceback
        print("Error al ver producto:")
        print(traceback.format_exc())
        flash(f'Error al cargar el producto: {str(e)}', 'danger')
        return redirect(url_for('products.index'))

@bp.route('/<int:product_id>/delete', methods=['POST'])
@login_required
def delete_product(product_id):
    """
    Enviar producto a la papelera en WooCommerce
    """
    try:
        from woocommerce import API
        from flask import current_app
        
        # Conectar con API de WooCommerce
        wcapi = API(
            url=current_app.config['WC_API_URL'],
            consumer_key=current_app.config['WC_CONSUMER_KEY'],
            consumer_secret=current_app.config['WC_CONSUMER_SECRET'],
            version="wc/v3",
            timeout=30
        )
        
        # Obtener el producto localmente para verificar tipo
        product = Product.query.get(product_id)
        if not product:
            return jsonify({'success': False, 'error': 'Producto no encontrado localmente'}), 404

        # Enviar a la papelera (force=False)
        response = wcapi.delete(f"products/{product_id}", params={"force": False})
        
        if response.status_code == 200:
            # Opcional: Actualizar estado local inmediatamente para evitar lag de sync
            try:
                product.post_status = 'trash'
                db.session.commit()
            except:
                db.session.rollback()
                
            return jsonify({
                'success': True, 
                'message': f'Producto "{product.post_title}" enviado a la papelera exitosamente.'
            })
        else:
            error_data = response.json()
            return jsonify({
                'success': False, 
                'error': error_data.get('message', 'Error en la API de WooCommerce')
            }), response.status_code
            
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/<int:product_id>/edit')
@login_required
def edit_product(product_id):
    """
    Formulario para editar un producto existente
    """
    try:
        product = Product.query.get(product_id)
        if not product:
            flash(f'Producto #{product_id} no encontrado.', 'danger')
            return redirect(url_for('products.index'))

        # OPTIMIZACIÓN: Pre-cargar metadatos e imágenes
        all_to_load = [product]
        variations = []
        if product.post_type == 'product':
            variations = Product.query.filter(
                Product.post_parent == product_id, 
                Product.post_type == 'product_variation',
                Product.post_status != 'trash'
            ).all()
            all_to_load.extend(variations)
        
        Product.preload_metadata_for_products(all_to_load)
        Product.preload_images_for_products(all_to_load)

        # Obtener galerías de imágenes (metadato _product_image_gallery)
        gallery_ids = product.get_meta('_product_image_gallery')
        gallery_urls = []
        if gallery_ids:
            ids = [int(id_str) for id_str in gallery_ids.split(',') if id_str]
            # Podríamos pre-cargar estas imágenes también si fuera necesario
            # Por ahora lo dejamos simple
            for img_id in ids:
                # Query simple para obtener URL de cada imagen de galería
                # (Idealmente optimizar esto después)
                meta = ProductMeta.query.filter_by(post_id=img_id, meta_key='_wp_attached_file').first()
                if meta:
                    base_url = 'https://www.izistoreperu.com/wp-content/uploads/'
                    gallery_urls.append(base_url + meta.meta_value)

        # Obtener IDs de taxonomías seleccionadas
        from sqlalchemy import text
        tax_query = text("""
            SELECT tt.term_id, tt.taxonomy
            FROM wpyz_term_relationships tr
            INNER JOIN wpyz_term_taxonomy tt ON tr.term_taxonomy_id = tt.term_taxonomy_id
            WHERE tr.object_id = :pid AND tt.taxonomy IN ('product_cat', 'product_tag', 'product_brand')
        """)
        results = db.session.execute(tax_query, {"pid": product_id})
        selected_tax = {'product_cat': [], 'product_tag': [], 'product_brand': []}
        for row in results:
            term_id, taxonomy = row[0], row[1]
            if taxonomy in selected_tax:
                selected_tax[taxonomy].append(term_id)

        return render_template(
            'products/edit.html', 
            product=product,
            variations=variations,
            gallery_urls=gallery_urls,
            selected_cats=selected_tax['product_cat'],
            selected_tags=selected_tax['product_tag'],
            selected_brands=selected_tax['product_brand'],
            title=f'Editar: {product.post_title[:30]}...'
        )
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        flash(f'Error al abrir editor: {str(e)}', 'danger')
        return redirect(url_for('products.index'))


@bp.route('/<int:product_id>/update', methods=['POST'])
@login_required
def update_product(product_id):
    """
    Actualizar un producto existente en WooCommerce mediante API REST
    """
    try:
        from woocommerce import API
        from flask import current_app
        
        data = request.get_json()
        
        # Conectar con API de WooCommerce
        wcapi = API(
            url=current_app.config['WC_API_URL'],
            consumer_key=current_app.config['WC_CONSUMER_KEY'],
            consumer_secret=current_app.config['WC_CONSUMER_SECRET'],
            version="wc/v3",
            timeout=60
        )
        
        # Construir objeto de actualización para producto padre
        update_data = {
            "name": data.get('title'),
            "status": data.get('status'),
            "description": data.get('description'),
            "short_description": data.get('short_description'),
            "sku": data.get('sku'),
            "regular_price": data.get('regular_price'),
            "sale_price": data.get('sale_price') if data.get('sale_price') else '',
            "manage_stock": True,
            "stock_quantity": int(data.get('stock_quantity', 0)) if data.get('stock_quantity') else 0,
            "stock_status": data.get('stock_status'),
            "weight": data.get('weight'),
            "dimensions": {
                "length": data.get('length'),
                "width": data.get('width'),
                "height": data.get('height')
            }
        }

        # Slug manual
        if data.get('slug'):
            update_data['slug'] = data.get('slug')

        # Imágenes (Principal + Galería)
        if data.get('image_url'):
            images = [{"src": data.get('image_url')}]
            if data.get('gallery_images'):
                for img_url in data.get('gallery_images'):
                    images.append({"src": img_url})
            update_data['images'] = images

        # Taxonomías
        if data.get('categories'):
            update_data['categories'] = [{"id": int(cat_id)} for cat_id in data.get('categories')]
        if data.get('tags'):
            update_data['tags'] = [{"id": int(tag_id)} for tag_id in data.get('tags')]
        # Nota: Marcas a veces requieren metadatos especiales dependiendo del plugin

        # Metadata SEO (Yoast)
        meta_data = []
        if data.get('seo_title'):
            meta_data.append({"key": "_yoast_wpseo_title", "value": data.get('seo_title')})
        if data.get('seo_description'):
            meta_data.append({"key": "_yoast_wpseo_metadesc", "value": data.get('seo_description')})
        if data.get('focus_keyphrase'):
            meta_data.append({"key": "_yoast_wpseo_focuskw", "value": data.get('focus_keyphrase')})
        
        if meta_data:
            update_data['meta_data'] = meta_data

        # Realizar actualización vía API
        response = wcapi.put(f"products/{product_id}", update_data)
        
        if response.status_code not in [200, 201]:
            error_msg = response.json().get('message', 'Error desconocido en WooCommerce API')
            return jsonify({'success': False, 'error': error_msg}), response.status_code

        # Manejo de marcas (Brands) vía SQL directo (consistente con store_product)
        brands = data.get('brands')
        if brands is not None:
            from sqlalchemy import text
            # Limpiar marcas actuales
            delete_query = text("""
                DELETE tr FROM wpyz_term_relationships tr
                INNER JOIN wpyz_term_taxonomy tt ON tr.term_taxonomy_id = tt.term_taxonomy_id
                WHERE tr.object_id = :pid AND tt.taxonomy = 'product_brand'
            """)
            db.session.execute(delete_query, {'pid': product_id})
            
            # Insertar nuevas
            for brand_id in brands:
                if brand_id:
                    insert_query = text("""
                        INSERT INTO wpyz_term_relationships (object_id, term_taxonomy_id)
                        SELECT :pid, term_taxonomy_id
                        FROM wpyz_term_taxonomy
                        WHERE term_id = :brand_id AND taxonomy = 'product_brand'
                        ON DUPLICATE KEY UPDATE term_taxonomy_id=term_taxonomy_id
                    """)
                    db.session.execute(insert_query, {'pid': product_id, 'brand_id': int(brand_id)})
            db.session.commit()

        # Actualizar variaciones si se enviaron
        if data.get('variations'):
            batch_data = {"update": []}
            for v in data.get('variations'):
                update_item = {
                    "id": v['id'],
                    "sku": v['sku'],
                    "regular_price": v['regular_price'],
                    "sale_price": v['sale_price'] if v['sale_price'] else '',
                    "stock_quantity": int(v['stock'] or 0),
                    "manage_stock": True
                }
                
                if v.get('image_url'):
                    update_item['image'] = {"src": v['image_url']}
                    
                batch_data["update"].append(update_item)
            
            if batch_data["update"]:
                wcapi.post(f"products/{product_id}/variations/batch", batch_data)

        return jsonify({
            'success': True, 
            'message': 'Producto actualizado correctamente en WooCommerce.'
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/create')
@login_required
def create():
    """
    Formulario para crear un nuevo producto

    URL: /products/create
    """
    return render_template('products/create.html', title='Crear Producto')


@bp.route('/create-variable-wizard')
@login_required
def create_variable_wizard():
    """
    Wizard para crear producto variable

    URL: /products/create-variable-wizard
    """
    return render_template('products/create_variable.html', title='Crear Producto Variable')


@bp.route('/create', methods=['POST'])
@login_required
def store_product():
    """
    Guardar nuevo producto en WooCommerce mediante API REST

    Campos esperados:
    - Información básica: title, slug, description, short_description, status
    - Precios: regular_price, sale_price, cost_price
    - Inventario: sku, stock_quantity, stock_status, low_stock_threshold, manage_stock
    - Peso y dimensiones: weight, length, width, height
    - Imágenes: image_url, gallery_images[]
    - Taxonomías: categories[], tags[], brands[]
    - SEO: seo_title, seo_description, focus_keyphrase, seo_keywords
    - Tipo: product_type (simple/variable)
    - Atributos (si es variable): attributes[]
    - Variaciones (si es variable): variations[]
    """
    try:
        from woocommerce import API
        from flask import current_app
        import re

        # Conectar con API de WooCommerce
        wcapi = API(
            url=current_app.config['WC_API_URL'],
            consumer_key=current_app.config['WC_CONSUMER_KEY'],
            consumer_secret=current_app.config['WC_CONSUMER_SECRET'],
            version="wc/v3",
            timeout=30
        )

        # Obtener datos del formulario
        data = request.get_json() if request.is_json else request.form.to_dict()

        # Construir objeto de producto para WooCommerce API
        product_data = {
            "name": data.get('title', '').strip(),
            "type": data.get('product_type', 'simple'),  # simple o variable
            "status": data.get('status', 'draft'),  # draft, publish, private
            "description": data.get('description', '').strip(),
            "short_description": data.get('short_description', '').strip(),
            "sku": data.get('sku', '').strip(),
            "regular_price": data.get('regular_price', ''),
            "sale_price": data.get('sale_price', '') if data.get('sale_price') else '',
            "manage_stock": data.get('manage_stock', 'yes') == 'yes',
            "stock_quantity": int(data.get('stock_quantity', 0)) if data.get('stock_quantity') else None,
            "stock_status": data.get('stock_status', 'instock'),
            "low_stock_amount": int(data.get('low_stock_threshold', '')) if data.get('low_stock_threshold') else None,
            "weight": data.get('weight', '').strip(),
            "dimensions": {
                "length": data.get('length', '').strip(),
                "width": data.get('width', '').strip(),
                "height": data.get('height', '').strip()
            }
        }

        # Generar slug automáticamente si no se proporcionó
        if data.get('slug'):
            product_data['slug'] = data.get('slug').strip()
        else:
            # Generar slug desde el título
            slug = re.sub(r'[^a-z0-9]+', '-', data.get('title', '').lower().strip())
            slug = slug.strip('-')
            product_data['slug'] = slug

        # Imagen principal
        if data.get('image_url'):
            product_data['images'] = [{"src": data.get('image_url').strip()}]

            # Agregar galería de imágenes si existen
            gallery_images = request.form.getlist('gallery_images[]') if not request.is_json else data.get('gallery_images', [])
            if gallery_images:
                for img_url in gallery_images:
                    if img_url and img_url.strip():
                        product_data['images'].append({"src": img_url.strip()})

        # Categorías
        categories = request.form.getlist('categories[]') if not request.is_json else data.get('categories', [])
        if categories:
            product_data['categories'] = [{"id": int(cat_id)} for cat_id in categories if cat_id]

        # Tags
        tags = request.form.getlist('tags[]') if not request.is_json else data.get('tags', [])
        if tags:
            product_data['tags'] = [{"id": int(tag_id)} for tag_id in tags if tag_id]

        # Brands (taxonomía personalizada)
        brands = request.form.getlist('brands[]') if not request.is_json else data.get('brands', [])

        # Meta data para campos personalizados
        meta_data = []

        # Precio de costo (campo personalizado)
        if data.get('cost_price'):
            meta_data.append({
                "key": "_cost_price",
                "value": data.get('cost_price').strip()
            })

        # Yoast SEO fields
        if data.get('seo_title'):
            meta_data.append({
                "key": "_yoast_wpseo_title",
                "value": data.get('seo_title').strip()
            })

        if data.get('seo_description'):
            meta_data.append({
                "key": "_yoast_wpseo_metadesc",
                "value": data.get('seo_description').strip()
            })

        if data.get('focus_keyphrase'):
            meta_data.append({
                "key": "_yoast_wpseo_focuskw",
                "value": data.get('focus_keyphrase').strip()
            })
            meta_data.append({
                "key": "_yoast_wpseo_focuskw_text_input",
                "value": data.get('focus_keyphrase').strip()
            })

        if data.get('seo_keywords'):
            meta_data.append({
                "key": "_yoast_wpseo_keywordsynonyms",
                "value": data.get('seo_keywords').strip()
            })

        if meta_data:
            product_data['meta_data'] = meta_data

        # Crear producto en WooCommerce
        response = wcapi.post("products", product_data)

        if response.status_code == 201:
            product_created = response.json()
            product_id = product_created['id']

            # Si tiene brands, agregar mediante API de términos
            if brands:
                from sqlalchemy import text
                for brand_id in brands:
                    if brand_id:
                        # Insertar relación en wpyz_term_relationships
                        query = text("""
                            INSERT INTO wpyz_term_relationships (object_id, term_taxonomy_id)
                            SELECT :product_id, term_taxonomy_id
                            FROM wpyz_term_taxonomy
                            WHERE term_id = :brand_id AND taxonomy = 'product_brand'
                            ON DUPLICATE KEY UPDATE term_taxonomy_id=term_taxonomy_id
                        """)
                        db.session.execute(query, {'product_id': product_id, 'brand_id': int(brand_id)})
                db.session.commit()

            return jsonify({
                'success': True,
                'message': 'Producto creado exitosamente',
                'product_id': product_id,
                'product': product_created
            }), 201
        else:
            error_message = response.json().get('message', 'Error desconocido')
            return jsonify({
                'success': False,
                'error': f'Error al crear producto: {error_message}',
                'status_code': response.status_code,
                'response': response.json()
            }), response.status_code

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/export-excel')
@login_required
def export_excel():
    """
    Exportar productos a Excel con filtros de fecha y título

    Parámetros GET:
    - date_from: Fecha desde (YYYY-MM-DD) - obligatorio
    - date_to: Fecha hasta (YYYY-MM-DD) - obligatorio
    - title: Filtro de título (opcional)

    Columnas del Excel:
    ID | Título | SKU | [Atributos dinámicos] | ID Padre | Descripción Corta |
    Descripción Larga | Precio Regular | Precio Oferta | URL Imagen | Stock
    """
    try:
        from datetime import datetime, timedelta
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from flask import make_response
        from io import BytesIO

        # Obtener parámetros
        date_from_str = request.args.get('date_from')
        date_to_str = request.args.get('date_to')
        title_filter = request.args.get('title', '').strip()

        # Validar parámetros obligatorios
        if not date_from_str or not date_to_str:
            return jsonify({
                'success': False,
                'error': 'Las fechas "Desde" y "Hasta" son obligatorias'
            }), 400

        # Convertir fechas
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
            # Agregar 23:59:59 a date_to para incluir todo el día
            date_to = date_to.replace(hour=23, minute=59, second=59)
        except ValueError as e:
            return jsonify({
                'success': False,
                'error': f'Formato de fecha inválido. Recibido: desde={date_from_str}, hasta={date_to_str}. Se esperaba formato YYYY-MM-DD (ej: 2026-01-26)'
            }), 400

        # PASO 1: Consultar productos del rango de fechas
        query = db.session.query(Product).filter(
            Product.post_type.in_(['product', 'product_variation']),
            Product.post_date >= date_from,
            Product.post_date <= date_to
        )

        # Filtrar por título si se especificó
        if title_filter:
            query = query.filter(Product.post_title.ilike(f'%{title_filter}%'))

        products = query.order_by(Product.post_date.desc()).all()

        if not products:
            return jsonify({
                'success': False,
                'error': f'No se encontraron productos entre {date_from_str} y {date_to_str}'
            }), 404

        # PASO 2: Obtener todos los IDs de productos
        product_ids = [p.ID for p in products]

        # PASO 2.5: Obtener IDs de productos padre para variantes
        parent_ids = set()
        for p in products:
            if p.post_type == 'product_variation' and p.post_parent > 0:
                parent_ids.add(p.post_parent)

        # Cargar productos padre que no están en la lista
        parent_products = {}
        if parent_ids:
            missing_parent_ids = parent_ids - set(product_ids)
            if missing_parent_ids:
                parents = db.session.query(Product).filter(
                    Product.ID.in_(list(missing_parent_ids))
                ).all()
                for parent in parents:
                    parent_products[parent.ID] = parent

        # PASO 3: Eager load de todos los metadatos
        meta_keys_base = ['_sku', '_price', '_regular_price', '_sale_price', '_stock',
                         '_stock_status', '_thumbnail_id', '_product_image_gallery', '_product_type']

        all_meta = db.session.query(
            ProductMeta.post_id,
            ProductMeta.meta_key,
            ProductMeta.meta_value
        ).filter(
            ProductMeta.post_id.in_(product_ids)
        ).all()

        # Crear diccionario de metadatos
        meta_dict = {}
        for post_id, meta_key, meta_value in all_meta:
            if post_id not in meta_dict:
                meta_dict[post_id] = {}
            meta_dict[post_id][meta_key] = meta_value

        # PASO 3.5: Obtener URLs de imágenes desde los thumbnail IDs
        thumbnail_ids = set()
        for post_id in product_ids:
            thumbnail_id = meta_dict.get(post_id, {}).get('_thumbnail_id')
            if thumbnail_id:
                try:
                    thumbnail_ids.add(int(thumbnail_id))
                except (ValueError, TypeError):
                    pass

        # Consultar rutas de archivos desde wpyz_postmeta
        image_urls = {}
        if thumbnail_ids:
            image_meta = db.session.query(
                ProductMeta.post_id,
                ProductMeta.meta_value
            ).filter(
                ProductMeta.post_id.in_(list(thumbnail_ids)),
                ProductMeta.meta_key == '_wp_attached_file'
            ).all()

            # Construir URLs completas
            base_url = 'https://www.izistoreperu.com/wp-content/uploads/'
            for post_id, file_path in image_meta:
                if file_path:
                    image_urls[str(post_id)] = base_url + file_path

        # PASO 4: Detectar todos los atributos únicos (para columnas dinámicas)
        # Atributos específicos que queremos detectar
        known_attributes = ['color', 'conector', 'talla', 'medidas', 'medida']
        all_attributes = set()

        for post_id in product_ids:
            product_meta = meta_dict.get(post_id, {})
            for key in product_meta.keys():
                # Buscar tanto attribute_pa_ como attribute_ (sin pa_)
                if key.startswith('attribute_'):
                    # Extraer nombre del atributo
                    if key.startswith('attribute_pa_'):
                        attr_slug = key.replace('attribute_pa_', '')
                    else:
                        attr_slug = key.replace('attribute_', '')
                    all_attributes.add(attr_slug)

        # Ordenar atributos alfabéticamente y capitalizar
        sorted_attributes = sorted(list(all_attributes))

        # Debug: Log de atributos detectados
        print(f"DEBUG: Atributos detectados: {sorted_attributes}")

        # PASO 4.5: Simplificado - ya no necesitamos consultar wpyz_terms
        # Usaremos los slugs directamente para mantener consistencia con los títulos

        # PASO 5: Crear archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"

        # PASO 6: Crear encabezados
        headers = ['ID', 'Título', 'SKU']
        # Capitalizar nombres de atributos para headers
        headers.extend([attr.replace('_', ' ').title() for attr in sorted_attributes])
        headers.extend(['ID Post', 'ID Padre', 'Descripción Corta', 'Descripción Larga',
                       'Precio Regular', 'Precio Oferta', 'URL Imagen', 'Stock'])

        # Escribir encabezados con estilo
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # PASO 7: Escribir datos de productos
        for row_num, product in enumerate(products, 2):
            product_meta = meta_dict.get(product.ID, {})

            col_num = 1

            # ID
            ws.cell(row=row_num, column=col_num, value=product.ID)
            col_num += 1

            # Título
            ws.cell(row=row_num, column=col_num, value=product.post_title)
            col_num += 1

            # SKU
            sku = product_meta.get('_sku', '')
            ws.cell(row=row_num, column=col_num, value=sku)
            col_num += 1

            # Atributos dinámicos
            for attr_slug in sorted_attributes:
                # Buscar el atributo con ambos formatos: attribute_pa_ y attribute_
                attr_key_pa = 'attribute_pa_' + attr_slug
                attr_key_no_pa = 'attribute_' + attr_slug

                attr_value_slug = product_meta.get(attr_key_pa, '') or product_meta.get(attr_key_no_pa, '')

                # Usar el slug directamente sin traducción
                # Esto mantiene consistencia con los títulos de productos
                attr_value = attr_value_slug

                ws.cell(row=row_num, column=col_num, value=attr_value)
                col_num += 1

            # ID Post - Muestra el ID del producto actual (todos los productos)
            ws.cell(row=row_num, column=col_num, value=product.ID)
            col_num += 1

            # ID Padre
            # Lógica simplificada:
            # - Variaciones: Mostrar ID del padre
            # - Productos padre (variables): Mostrar vacío
            # - Productos simples: Mostrar vacío
            if product.post_type == 'product_variation' and product.post_parent > 0:
                parent_id = product.post_parent
            else:
                parent_id = ''

            ws.cell(row=row_num, column=col_num, value=parent_id)
            col_num += 1

            # Descripción Corta y Larga
            # Para variaciones, SIEMPRE obtener del producto padre
            # Las variaciones suelen tener atributos en post_excerpt, no la descripción real
            if product.post_type == 'product_variation' and product.post_parent > 0:
                # Buscar padre en la lista de productos o en parent_products
                padre = next((p for p in products if p.ID == product.post_parent), None)
                if not padre and product.post_parent in parent_products:
                    padre = parent_products[product.post_parent]

                if padre:
                    desc_corta = padre.post_excerpt or ''
                    desc_larga = padre.post_content or ''
                else:
                    # Si no se encuentra el padre, usar vacío (no los atributos de la variación)
                    desc_corta = ''
                    desc_larga = ''
            else:
                # Para productos simples y variables (padres), usar su propia descripción
                desc_corta = product.post_excerpt or ''
                desc_larga = product.post_content or ''

            ws.cell(row=row_num, column=col_num, value=desc_corta)
            col_num += 1

            ws.cell(row=row_num, column=col_num, value=desc_larga)
            col_num += 1

            # Precio Regular
            regular_price = product_meta.get('_regular_price', '')
            ws.cell(row=row_num, column=col_num, value=regular_price)
            col_num += 1

            # Precio Oferta
            sale_price = product_meta.get('_sale_price', '')
            ws.cell(row=row_num, column=col_num, value=sale_price)
            col_num += 1

            # URL Imagen
            thumbnail_id = product_meta.get('_thumbnail_id', '')
            image_url = ''
            if thumbnail_id:
                image_url = image_urls.get(str(thumbnail_id), '')
            ws.cell(row=row_num, column=col_num, value=image_url)
            col_num += 1

            # Stock
            stock = product_meta.get('_stock', '')
            ws.cell(row=row_num, column=col_num, value=stock)

        # PASO 8: Ajustar anchos de columnas
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(headers[col_num - 1]))

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                   min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

            # Limitar ancho máximo
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # PASO 9: Generar archivo en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # PASO 10: Crear respuesta HTTP
        filename = f'productos_{date_from_str}_{date_to_str}.xlsx'
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'

        return response

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@bp.route('/get-categories')
@login_required
def get_categories():
    """
    Obtener todas las categorías de productos
    URL: /products/get-categories
    """
    try:
        from sqlalchemy import text
        query = text("""
            SELECT t.term_id, t.name, t.slug, tt.parent
            FROM wpyz_terms t
            INNER JOIN wpyz_term_taxonomy tt ON t.term_id = tt.term_id
            WHERE tt.taxonomy = 'product_cat'
            ORDER BY t.name ASC
        """)
        result = db.session.execute(query)
        categories = []
        for row in result:
            categories.append({'id': row[0], 'name': row[1], 'slug': row[2], 'parent_id': row[3]})
        return jsonify({'success': True, 'categories': categories})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/get-tags')
@login_required
def get_tags():
    """
    Obtener todos los tags de productos
    URL: /products/get-tags
    """
    try:
        from sqlalchemy import text
        query = text("""
            SELECT t.term_id, t.name, t.slug
            FROM wpyz_terms t
            INNER JOIN wpyz_term_taxonomy tt ON t.term_id = tt.term_id
            WHERE tt.taxonomy = 'product_tag'
            ORDER BY t.name ASC
        """)
        result = db.session.execute(query)
        tags = []
        for row in result:
            tags.append({'id': row[0], 'name': row[1], 'slug': row[2]})
        return jsonify({'success': True, 'tags': tags})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/get-brands')
@login_required
def get_brands():
    """
    Obtener todas las marcas de productos
    URL: /products/get-brands
    """
    try:
        from sqlalchemy import text
        query = text("""
            SELECT t.term_id, t.name, t.slug
            FROM wpyz_terms t
            INNER JOIN wpyz_term_taxonomy tt ON t.term_id = tt.term_id
            WHERE tt.taxonomy = 'product_brand'
            ORDER BY t.name ASC
        """)
        result = db.session.execute(query)
        brands = []
        for row in result:
            brands.append({'id': row[0], 'name': row[1], 'slug': row[2]})
        return jsonify({'success': True, 'brands': brands})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/get-attributes')
@login_required
def get_attributes():
    """
    Obtener todos los atributos de productos globales
    URL: /products/get-attributes
    """
    try:
        from woocommerce import API
        from flask import current_app

        wcapi = API(
            url=current_app.config['WC_API_URL'],
            consumer_key=current_app.config['WC_CONSUMER_KEY'],
            consumer_secret=current_app.config['WC_CONSUMER_SECRET'],
            version="wc/v3",
            timeout=30
        )

        response = wcapi.get("products/attributes")

        if response.status_code == 200:
            attributes = response.json()
            return jsonify({'success': True, 'attributes': attributes})
        else:
            return jsonify({'success': False, 'error': 'Error al obtener atributos'}), 500

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/get-attribute-terms/<int:attribute_id>')
@login_required
def get_attribute_terms(attribute_id):
    """
    Obtener términos de un atributo específico
    URL: /products/get-attribute-terms/1
    """
    try:
        from woocommerce import API
        from flask import current_app

        wcapi = API(
            url=current_app.config['WC_API_URL'],
            consumer_key=current_app.config['WC_CONSUMER_KEY'],
            consumer_secret=current_app.config['WC_CONSUMER_SECRET'],
            version="wc/v3",
            timeout=30
        )

        response = wcapi.get(f"products/attributes/{attribute_id}/terms", params={"per_page": 100})

        if response.status_code == 200:
            terms = response.json()
            return jsonify({'success': True, 'terms': terms})
        else:
            return jsonify({'success': False, 'error': 'Error al obtener términos'}), 500

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/create-variable', methods=['POST'])
@login_required
def create_variable_product():
    """
    Crear producto variable con todas sus variaciones

    Recibe:
    - Información del producto padre (paso 1)
    - Atributos y términos seleccionados (paso 2)
    - Variaciones generadas con precios/stock (paso 3)
    """
    try:
        from woocommerce import API
        from flask import current_app
        import re
        import itertools

        wcapi = API(
            url=current_app.config['WC_API_URL'],
            consumer_key=current_app.config['WC_CONSUMER_KEY'],
            consumer_secret=current_app.config['WC_CONSUMER_SECRET'],
            version="wc/v3",
            timeout=60
        )

        data = request.get_json()

        # PASO 1: Crear producto padre (variable)
        product_data = {
            "name": data.get('title', '').strip(),
            "type": "variable",
            "status": data.get('status', 'draft'),
            "description": data.get('description', '').strip(),
            "short_description": data.get('short_description', '').strip(),
            "weight": data.get('weight', '').strip(),
            "dimensions": {
                "length": data.get('length', '').strip(),
                "width": data.get('width', '').strip(),
                "height": data.get('height', '').strip()
            }
        }

        # Slug
        if data.get('slug'):
            product_data['slug'] = data.get('slug').strip()
        else:
            slug = re.sub(r'[^a-z0-9]+', '-', data.get('title', '').lower().strip())
            product_data['slug'] = slug.strip('-')

        # Imagen principal
        if data.get('image_url'):
            product_data['images'] = [{"src": data.get('image_url').strip()}]
            gallery_images = data.get('gallery_images', [])
            if gallery_images:
                for img_url in gallery_images:
                    if img_url and img_url.strip():
                        product_data['images'].append({"src": img_url.strip()})

        # Categorías
        categories = data.get('categories', [])
        if categories:
            product_data['categories'] = [{"id": int(cat_id)} for cat_id in categories if cat_id]

        # Tags
        tags = data.get('tags', [])
        if tags:
            product_data['tags'] = [{"id": int(tag_id)} for tag_id in tags if tag_id]

        # Brands
        brands = data.get('brands', [])

        # Atributos del producto
        attributes_data = []
        for attr in data.get('attributes', []):
            attributes_data.append({
                "id": attr['id'],
                "name": attr['name'],
                "visible": True,
                "variation": True,
                "options": attr['terms']
            })

        product_data['attributes'] = attributes_data

        # Meta data (SEO, costo)
        meta_data = []
        if data.get('cost_price'):
            meta_data.append({"key": "_cost_price", "value": data.get('cost_price').strip()})
        if data.get('seo_title'):
            meta_data.append({"key": "_yoast_wpseo_title", "value": data.get('seo_title').strip()})
        if data.get('seo_description'):
            meta_data.append({"key": "_yoast_wpseo_metadesc", "value": data.get('seo_description').strip()})
        if data.get('focus_keyphrase'):
            meta_data.append({"key": "_yoast_wpseo_focuskw", "value": data.get('focus_keyphrase').strip()})
            meta_data.append({"key": "_yoast_wpseo_focuskw_text_input", "value": data.get('focus_keyphrase').strip()})
        if data.get('seo_keywords'):
            meta_data.append({"key": "_yoast_wpseo_keywordsynonyms", "value": data.get('seo_keywords').strip()})

        if meta_data:
            product_data['meta_data'] = meta_data

        # Crear producto padre
        response = wcapi.post("products", product_data)

        if response.status_code != 201:
            error_message = response.json().get('message', 'Error desconocido')
            return jsonify({
                'success': False,
                'error': f'Error al crear producto: {error_message}',
                'response': response.json()
            }), response.status_code

        product_created = response.json()
        product_id = product_created['id']

        # Agregar brands si existen
        if brands:
            from sqlalchemy import text
            for brand_id in brands:
                if brand_id:
                    query = text("""
                        INSERT INTO wpyz_term_relationships (object_id, term_taxonomy_id)
                        SELECT :product_id, term_taxonomy_id
                        FROM wpyz_term_taxonomy
                        WHERE term_id = :brand_id AND taxonomy = 'product_brand'
                        ON DUPLICATE KEY UPDATE term_taxonomy_id=term_taxonomy_id
                    """)
                    db.session.execute(query, {'product_id': product_id, 'brand_id': int(brand_id)})
            db.session.commit()

        # PASO 2: Crear variaciones
        variations = data.get('variations', [])
        created_variations = []
        failed_variations = []

        for variation in variations:
            variation_data = {
                "regular_price": variation.get('regular_price', ''),
                "sale_price": variation.get('sale_price', '') if variation.get('sale_price') else '',
                "sku": variation.get('sku', '').strip(),
                "manage_stock": variation.get('manage_stock', True),
                "stock_quantity": int(variation.get('stock_quantity', 0)) if variation.get('stock_quantity') else 0,
                "stock_status": variation.get('stock_status', 'instock'),
                "attributes": []
            }

            # Atributos de la variación
            for attr_slug, term_slug in variation.get('attributes', {}).items():
                variation_data['attributes'].append({
                    "id": int(attr_slug.replace('attr_', '')),
                    "option": term_slug
                })

            # Imagen de variación
            if variation.get('image_url'):
                variation_data['image'] = {"src": variation.get('image_url').strip()}

            # Precio de costo de variación
            if variation.get('cost_price'):
                variation_data['meta_data'] = [{"key": "_cost_price", "value": variation.get('cost_price').strip()}]

            # Crear variación
            var_response = wcapi.post(f"products/{product_id}/variations", variation_data)

            if var_response.status_code == 201:
                created_variations.append(var_response.json())
            else:
                failed_variations.append({
                    'variation': variation,
                    'error': var_response.json()
                })

        return jsonify({
            'success': True,
            'message': 'Producto variable creado exitosamente',
            'product_id': product_id,
            'product': product_created,
            'variations_created': len(created_variations),
            'variations_failed': len(failed_variations),
            'failed_details': failed_variations
        }), 201

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500
